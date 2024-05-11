import openpyxl
from openpyxl.styles  import Font
import time
import os
import configparser
from calendar import month_name


class Transaction_book_calculator():
    # Class variables for configuration
    config = configparser.ConfigParser()
    setting_ini_path = f"{(os.path.dirname(os.path.abspath(__file__)))}\\setting.ini"
    config.read(setting_ini_path)

    def __init__(self, Transaction_detalies_list, transaction_name, add_or_sub, What_to_do,*args):
        #for Gui send Notification
        try:
            self.Notification_func=args[0]["Func"]
            self.Revert_all=args[0]["Revert"]
        except:
            self.Notification_func= self.dialHere
            self.Revert_all=self.dialHere
        
        if What_to_do=="trans":    
            self.Transaction_detalies_list = Transaction_detalies_list
            self.transaction_name = transaction_name
            self.add_or_sub = add_or_sub
            self.What_to_do = What_to_do
            self.Trans_variable_initialize()
            self.write_trans_excel_data()
        elif What_to_do == "Loan":
            self.Transaction_detalies_list = Transaction_detalies_list
            self.transaction_name = transaction_name
            self.add_or_sub = add_or_sub
            self.What_to_do = What_to_do
            self.Loan_Variable_initialize()

    def Trans_variable_initialize(self):
        self.solid_trans_name = "_".join(self.transaction_name.split(" "))
        try :
            self.directory_path = self.config.get("FILE_LOCATION", "location")
        except :
            newconfig= configparser.ConfigParser()
            newconfig.read(self.setting_ini_path)
            with open(self.setting_ini_path, "w") as f:
                newconfig.write(f)
            self.config.read(self.setting_ini_path)
            self.directory_path = self.config.get("FILE_LOCATION", "location")

        self.File_path = f"{self.directory_path}/{self.transaction_name}/{self.transaction_name}'s tranjaction of {time.strftime('%Y_%m')}.xlsx"
        last_month = int(self.config.get(self.solid_trans_name, f"{self.solid_trans_name.lower()}_last_month"))
        current_month = time.strftime('%m')

        if last_month != int(current_month):
            self.create_new_Month_file()
 
        self.last_saved_num = int(self.config.get(self.solid_trans_name, f"{self.solid_trans_name.lower()}_code"))
        self.amount = float(self.config.get(self.solid_trans_name, f"{self.solid_trans_name}_amounts"))

    def create_new_Month_file(self):    
        Wb = openpyxl.Workbook()
        ws = Wb.active
        ws.title = self.transaction_name
        ws.append(["SN", "Date", "Keywords", "Reason", "Income", "Expences"])
        # previous_date
        Month_name= int(self.config.get(self.solid_trans_name, f"{self.solid_trans_name.lower()}_last_month"))
        amt= float(self.config.get(self.solid_trans_name, f"{self.solid_trans_name.lower()}_amounts"))
        ws.append([">>",f"Month:{Month_name}","Last Month Money","Last Month Money",amt ,0])

        row = (tuple(ws.rows))
        for each_column in row[0]:
            each_column.font = Font(bold=True)
        try:    
            Wb.save(f"{self.directory_path}/{self.transaction_name}/{self.transaction_name}'s tranjaction of {time.strftime('%Y_%m')}.xlsx")
        except PermissionError:
            self.Notification_func(text="<<File Access Permission Error>>",color="#A94438")    
        except Exception as error_reason:
            self.Notification_func(text=error_reason,color="#A94438")    
        self.config.set(self.solid_trans_name, f"{self.solid_trans_name.lower()}_last_month", time.strftime('%m'))
        self.config.set(self.solid_trans_name, f"{self.solid_trans_name.lower()}_code", "0")

    def write_trans_excel_data(self):
        try:
            Wb = openpyxl.load_workbook(self.File_path)
            ws = Wb.active
            for each_transation in self.Transaction_detalies_list:
                self.last_saved_num += 1
                if self.add_or_sub == "ADD":
                    self.Incomes = each_transation["amount"]
                    self.Expences = 0
                    self.amount += float(self.Incomes)
                else:
                    self.Incomes = 0
                    self.Expences = each_transation["amount"]
                    self.amount -= float(self.Expences)

                current_time = time.strftime("%d, %H:%M:%S")
                ws.append([self.last_saved_num,
                        current_time,
                        ",".join(each_transation["keyword_collection"]),
                        each_transation["source_reason"],
                        self.Incomes, self.Expences])
                
            Wb.save(self.File_path)
            self.Revert_all()

            # Update the configuration values

            self.config.set(self.solid_trans_name, f"{self.solid_trans_name}_amounts", str(self.amount))
            self.config.set(self.solid_trans_name, f"{self.solid_trans_name}_code", str(self.last_saved_num))

            # Write the configuration changes
            with open(self.setting_ini_path, "w") as f:
                self.config.write(f)
            self.Notification_func(f"Saved in {self.transaction_name} File, SuccessFully","Green")    
        except PermissionError:
            self.Notification_func(text="<<File Access Permission Error>>",color="#A94438")
        except Exception as error_reason:
                self.Notification_func(text=error_reason,color="#A94438")     

    def dialHere(self, *args,**kwargs):
        try: 
            if args[0][0:5] =='Saved':
                print("Saved File, SuccessFully")
        except:pass     


    def Loan_Variable_initialize(self):
        try :
            self.directory_path_= f"{self.config.get('FILE_LOCATION', 'location')}/Loan/"
        except :
            newconfig= configparser.ConfigParser()
            newconfig.read(self.setting_ini_path)
            with open(self.setting_ini_path, "w") as f:
                newconfig.write(f)
            self.config.read(self.setting_ini_path)
            self.directory_path_= f"{self.config.get('FILE_LOCATION', 'location')}/Loan/"

        All_details_Dict ={}
        for each_trans in self.Transaction_detalies_list:
            Tangent_name= "_".join(each_trans['Tangent_Name'][0].split(" "))
            try:
                self.config.get("Loan", f"{Tangent_name}_amt")
                location= f"{self.directory_path_}Loan_with_{Tangent_name}.xlsx"
                print(Tangent_name)
                if Tangent_name in All_details_Dict:

                    All_details_Dict[Tangent_name].append([location,each_trans])
                else:    
                    All_details_Dict.update({Tangent_name:[[location,each_trans]]})

            except Exception as e:
                loaction= self.create_new_tangent_file(Tangent_name)
                All_details_Dict.update({Tangent_name:[[loaction,each_trans]]})
     
        self.write_Loan_tangent_file(All_details_Dict)       

    def create_new_tangent_file(self,tangent_name):
        Wb = openpyxl.Workbook()
        ws = Wb.active
        ws.title = tangent_name
        ws.append(["SN", "Date","Due Date", "Reason", "Loan Given", "Loan Collected", "Deposit"])
        row = (tuple(ws.rows))
        for each_column in row[0]:
            each_column.font = Font(bold=True)
        try:    
            Wb.save(f"{self.directory_path_}Loan_with_{tangent_name}.xlsx")
            self.config.set("Loan", f"{tangent_name.lower()}_amt", '0.0')
            self.config.set("Loan", f"{tangent_name.lower()}_code", '0')
            with open(self.setting_ini_path, "w") as f:
                self.config.write(f)
            location= f"{self.directory_path_}Loan_with_{tangent_name}.xlsx"
            return (location)

        except PermissionError:
            self.Notification_func(text="<<File Access Permission Error>>",color="#A94438")    
        except Exception as error_reason:
            print("error in create")
            self.Notification_func(text=error_reason,color="#A94438")       

    def write_Loan_tangent_file(self, All_Details_list):
        for each_tangent, details_List in All_Details_list.items():
            location= details_List[0][0]
            Tangent_Name= each_tangent
            try:
                Wb = openpyxl.load_workbook(location)
                ws = Wb.active    
                for each_loan_made in details_List:
                    loan_Details= each_loan_made[1]
                    due_date= loan_Details['Due_date']
                    source_reason= loan_Details['source_reason']
                    amount= float(loan_Details['amount'])
                    deposit_normal= loan_Details['deposit_normal']
                    if deposit_normal =="N":
                        deposit_normal=None
                    else:
                        deposit_normal="Paid Back"    
                    configured_amt=float(self.config.get("Loan", f"{Tangent_Name}_amt"))
                    configured_code=int(self.config.get("Loan", f"{Tangent_Name}_code"))
                    self.add_or_sub
                    configured_code += 1
                    if self.add_or_sub == "Take":
                        Loan_Collected = amount
                        loan_Given = 0
                        configured_amt -= (amount)
                    else:
                        Loan_Collected = 0
                        loan_Given = amount
                        configured_amt += (amount)

                    current_time = time.strftime("%Y-%m-%d, %H")
                    ws.append([configured_code,
                                current_time,
                                due_date,
                                source_reason,
                                loan_Given,
                                Loan_Collected,
                                deposit_normal])
                    
                    self.config.set("Loan", f"{Tangent_Name.lower()}_amt", str(configured_amt))
                    self.config.set("Loan", f"{Tangent_Name.lower()}_code", str(configured_code))
                    with open(self.setting_ini_path, "w") as f:
                        self.config.write(f)
                    Wb.save(location)
                    self.Revert_all()
                    if self.Notification_func !=0:
                        self.Notification_func(f"Saved in {self.transaction_name} File, SuccessFully","Green") 
            except PermissionError:
                self.Notification_func(text="<<File Access Permission Error>>",color="#A94438") 
            except Exception as error_reason:
                self.Notification_func(text=error_reason,color="#A94438") 
