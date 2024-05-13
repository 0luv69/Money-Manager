from tkinter import ttk
import tkinter as tk
from tkcalendar import DateEntry, Calendar
from tkinter import messagebox
import customtkinter as ctk
import time
from datetime import datetime, timedelta
from calendar import month_name
import os
import configparser
import openpyxl
import math
from transation_calculator import *

class Window(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.geometry("500x550")
        self.title("Money Manager")
        self.resizable(False,False)
        self.init_variable()

        WinBackEnd(self)
        self.mainloop()    

    def init_variable(self):
        self.setting_ini_path=f"{(os.path.dirname(os.path.abspath(__file__)))}\\setting.ini"
        
        self.current_date_var= ctk.StringVar()
        self.next_reminder_var= ctk.StringVar(value="2 PM")
        self.time_left_var= ctk.IntVar(value=5)
        self.User_Name=ctk.StringVar()
     
class WinBackEnd():
    def __init__(self,window):                                                                                                                                              
        self.window=window
        Setup_result=Backend_Inital_page(self.window)
        if Setup_result.Setup_result:
            Backend_First_page(self.window)
            
class Backend_Inital_page():
    def __init__(self,window):   
        self.window=window  
        self.Money_Element_list=[]
        self.config= configparser.ConfigParser()    
        self.config.read(self.window.setting_ini_path)
        try:
            self.setup_user_name=self.config.get("FILE_LOCATION","setup_user_name")  
        except:
            self.config.add_section('FILE_LOCATION')
            self.config.set('FILE_LOCATION', 'location', '')
            self.config.set('FILE_LOCATION', 'setup_user_name', '')
            self.config.set('FILE_LOCATION', 'elements_for_money_track', '')

            self.config.add_section('Keywords')
            self.config.set('Keywords','keywords','')

            self.config.add_section("Loan")
            self.config.set("Loan", "tangents","")

            with open(self.window.setting_ini_path, 'w') as config_file:
                self.config.write(config_file)
            self.config= configparser.ConfigParser()    
            self.config.read(self.window.setting_ini_path)
            self.setup_user_name=self.config.get("FILE_LOCATION","setup_user_name")

        self.directory_path=self.config.get("FILE_LOCATION","location")

        if not self.setup_user_name:
            self.directory_path="C:/Users/baniy/Documents"
            self.inital_page=Inital_page(self.window,self)
            self.Setup_result=False
        else:
            self.Setup_result=True
 
    def database_folder_selector(self):
        # setting up the location 
        self.directory_path=ctk.filedialog.askdirectory(title="Select Storage",initialdir=f"{os.path.expanduser('~')}\\Documents",mustexist=True)
        if self.directory_path :
            self.inital_page.selection_btn.configure(text="Selected ",
                                                     border_color="#D2DE32",
                                                     fg_color="green",
                                                     border_width=3, 
                                                     )
        elif not self.directory_path:
            self.directory_pat=""
            self.inital_page.selection_btn.configure(text="Select",
                                                     border_color="white",
                                                     fg_color="#4F6F52",
                                                     border_width=1, 
                                                     )
            Notification(self.window,f"Plz select the folder to store Data !!!!!", "red",6)

    def money_track_adder(self,*args):
        if self.element_name.get() != "":
            return_value =self.Save_changes_in_element_name()
            if return_value==0:
                return
        if self.element_name.get() == "New Element":
            Notification(text="First Edit Inital Name of Element",window=self.window)    
        else:
            self.button_structure("New Element",self.inital_page,len(self.Money_Element_list))
            self.selection(len(self.Money_Element_list)-1,self.inital_page)

    def money_track_remover(self):
        self.button_instance.destroy()
        self.Money_Element_list[self.index_of_selected_element]="removed"
        self.element_name.delete(0,ctk.END)

    def Save_changes_in_element_name(self,*args): 
        if self.button_instance.cget("text") != self.element_name.get():
            if len(self.element_name.get())>20:
                Notification(text="Too long name reduce it, (max:20)",window=self.window)
                return 0
            elif len(self.element_name.get()) <3:
                Notification(text="Too Short name Increase it, (min:3)",window=self.window)
                return 0
            elif any(char in self.element_name.get() for char in['<', '>', ':', '"', '/', '\\', '|', '?', '*','.']):
                Notification(text="PLz dont keep character like <, >, :, \", /, \\, |, ?, *, . and others",window=self.window)
                return 0
            else:    
                self.button_instance.configure(text=self.element_name.get())
                return 1

    def selection(self,index_of_element,inital_page_self):
        try:
            if self.element_name.get() != "":
                self.Save_changes_in_element_name()
        except AttributeError:pass    

        self.index_of_selected_element= index_of_element
        for i in range(len(self.Money_Element_list)):
            if self.Money_Element_list[i] == "removed":
                pass
            else:
                self.Money_Element_list[i].configure(border_width=0,text_color="White",fg_color="#739072")

        self.button_instance=self.Money_Element_list[index_of_element]
        self.button_instance.configure(border_color="RED", border_width=2,text_color="black",fg_color="green")

        self.element_name=inital_page_self.elements_name
        self.element_name.delete(0,ctk.END)
        self.element_name.insert(ctk.END, self.button_instance.cget("text"))

    def button_structure(self,text,inital_page_self,index_of_element):
        new_btn=ctk.CTkButton(inital_page_self.money_track_adder_frame, 
                      text=text,
                      fg_color="#739072",
                      hover_color="green",
                      cursor="hand2",
                      command=lambda :self.selection(index_of_element,inital_page_self)
                      )
        new_btn.pack(side="right",padx=5)
        self.Money_Element_list.append(new_btn)

    def Go_checker(self,*args):        
        if not self.directory_path:
            Notification(self.window,f"Plz select the \"Folder\" !!!", "red",6)
        elif not self.inital_page.Name_user.get():
            Notification(self.window,f"Plz Fill the \"Name\" !!!", "red",6)    
        elif self.index_of_selected_element<1:
            Notification(self.window,f"Plz provide atleast 2 Money Tracking Element !!!", "red",6)    
        else:
            self.main_folder_name=f"{self.directory_path}/Money Manager"
            i=0
            while(os.path.exists(self.main_folder_name)):
                i+=1
                self.main_folder_name= f"{self.directory_path}/Money Manager{i}"
            self.config.set("FILE_LOCATION", "setup_user_name", self.inital_page.Name_user.get())  
            self.config.set("FILE_LOCATION", "location", self.main_folder_name) 
            self.elements_for_money_track=[]
            for  element in self.Money_Element_list:
                if element =="removed":
                    pass
                else:
                    self.elements_for_money_track.append(element.cget("text"))   
            self.config.set("FILE_LOCATION","elements_for_money_track",','.join(self.elements_for_money_track))        
            self.config.set("Keywords","Keywords",','.join(self.elements_for_money_track))        
            
            for Each_element in self.elements_for_money_track:
                Each_element= "_".join(Each_element.split(" "))
                self.config[Each_element]={
                    f"{Each_element}_amounts":0,
                    f"{Each_element}_last_month":0,
                    f"{Each_element}_code":0,
                }

            with open(self.window.setting_ini_path,"w") as f:
                self.config.write(f)
            self.inital_page.destroy()   
            self.create_files()
            Backend_First_page(self.window) 

    def create_files(self):
        current_date=time.strftime("%Y_%m")
        os.makedirs(self.main_folder_name, exist_ok=True)
        for each_name in  self.elements_for_money_track:
            Each_element_folder=f"{self.main_folder_name}/{each_name}"
            os.makedirs(Each_element_folder,exist_ok=True)
            wb=openpyxl.Workbook()
            ws=wb.active
            ws.title=each_name
            ws.append(["SN","Date","Keywords","Reason","Income","Expences"])
            row=(tuple(ws.rows))
            for each_column in row[0]:
                each_column.font=Font(bold=True)
            wb.save(f"{Each_element_folder}\\{each_name}'s tranjaction of {current_date}.xlsx")

        # for rent file
        os.makedirs(f"{self.main_folder_name}/Loan",exist_ok=True) 
 
class Inital_page(ctk.CTkFrame):
    def __init__(self,window,Backend_0th_page,):
        super().__init__(window)
        self.pack(fill="both",expand=True)
        self.window= window
        
        ctk.CTkLabel(self,text="Money Manager", 
                     font=ctk.CTkFont("Showcard Gothic Regular", 40, "bold")
                     ).place(relx=.15,rely=.02,relwidth=.7)
         
        selection_frame= ctk.CTkFrame(self,fg_color="#D2E3C8")
        selection_frame.place(relx= .08, rely = .14 ,relwidth= .84,relheight=.8)

        ctk.CTkLabel(selection_frame,text="Setup Page", 
                     text_color="#4F6F52",
                     font=ctk.CTkFont("Agency fb", 30, "bold")
                     ).place(relx=.15,rely=0,relwidth=.7)

        selection_inner_frame= ctk.CTkFrame(selection_frame,fg_color="#739072",corner_radius=10)
        selection_inner_frame.place(relx= .08, rely = .1 ,relwidth= .84,relheight=.84)

        ctk.CTkLabel(selection_inner_frame, 
                     text_color="#000000",
                     text="Select the folder to save your Data",
                     font=ctk.CTkFont("Agency fb", 25,)
                     ).place(relx=.02,rely=.04,relwidth=.96,)
        
        #select button
        self.selection_btn=ctk.CTkButton(selection_inner_frame,
                      cursor="hand2",
                      hover_color="green",
                      text="Selected Document Path",
                      border_color="#D2DE32",
                      fg_color="green",
                      border_width=3, 
                      command=Backend_0th_page.database_folder_selector,
                      )
        self.selection_btn.place(relx=.3,rely=.16,relheight=.1)

        #name label
        ctk.CTkLabel(selection_inner_frame,
                     text="Name: ",
                     text_color="#000000",
                     font=ctk.CTkFont("cambria math", 20,)
                     ).place(relx=.06,rely=.2)
        
        self.Name_user=ctk.CTkEntry(selection_inner_frame,
                     fg_color="white",
                     text_color="black",
                     corner_radius=10
                     )
        self.Name_user.place(relx= .25,rely=.315,relwidth=.68)

        ctk.CTkLabel(selection_inner_frame,
                     text_color="#000000",
                     text="ADD Element to Track Money on :",
                     font=ctk.CTkFont("Agency fb", 22,)
                     ).place(relx=.02,rely=.43,relwidth=.96)
           
        self.elements_name=ctk.CTkEntry(selection_inner_frame,fg_color="white",text_color="black")
        self.elements_name.place(relx=.16,rely=.54,relwidth=.4)

        ctk.CTkButton(selection_inner_frame,text="➕",
                        fg_color="Dark green",
                        hover_color="#4F6F52",
                        cursor="hand2",
                        command=Backend_0th_page.money_track_adder ,                       
                        ).place(relx=.58,rely=.54,relwidth=.1)  
          
        ctk.CTkButton(selection_inner_frame,text="➖",
                        fg_color="red",
                        hover_color="#FF0060",
                        cursor="hand2",
                        command=Backend_0th_page.money_track_remover,
                        ).place(relx=.7,rely=.54,relwidth=.1)

        ctk.CTkButton(selection_inner_frame,text="✔",
                        fg_color="blue",
                        hover_color="#711DB0",
                        cursor="hand2",
                        command=Backend_0th_page.Save_changes_in_element_name,
                        ).place(relx=.82,rely=.538,relwidth=.1)

        self.money_track_adder_frame=ctk.CTkScrollableFrame(selection_inner_frame,
                                                       orientation="horizontal",
                                                       fg_color="#4F6F52",
                                                       scrollbar_button_color="black",
                                                       scrollbar_button_hover_color="#010A43")
        
        self.money_track_adder_frame.place(relx=.04,rely=.64,relwidth=.92,relheight=.23)
        
        #button structure
        Backend_0th_page.button_structure("Hand Cash",self,0)
        Backend_0th_page.selection(0,self)

        ctk.CTkButton(selection_inner_frame,
                      text="Go..",
                      fg_color="#4F6F52",
                      cursor="star",
                      hover_color="green",
                      border_color="yellow",
                      border_width=1,
                      command=Backend_0th_page.Go_checker
                      ).place(relx= .4,rely=.88,relheight=.1,relwidth=.2)

        #binding.........
        self.window.bind("<Control-s>",Backend_0th_page.Go_checker)
        self.elements_name.bind("<Return>",Backend_0th_page.money_track_adder)

class Backend_First_page():
    def __init__(self,window):
        self.window=window
        self.config=configparser.ConfigParser()
        self.config.read(self.window.setting_ini_path)

        self.elements=(self.config.get("FILE_LOCATION","elements_for_money_track")).split(",")
        self.window.User_Name.set(self.config.get("FILE_LOCATION","setup_user_name"))
        self.change_datetime()

        self.first_page=First_page(self.window)
        self.Button_structure(("Check","Loan"))
        for i in range(0,len(self.elements),2):
            if i== len(self.elements)-1:
                self.Button_structure([self.elements[i]],side="bottom")
            else:
                self.Button_structure((self.elements[i],self.elements[i+1]))
  
    def change_datetime(self):
        current_time_struct = time.localtime()
        formatted_time = time.strftime("%Y/%m/%d, %I:%M %p", current_time_struct)
        self.window.current_date_var.set(formatted_time)

    def Button_structure(self,texts,side="left"):
        button_frame=ctk.CTkFrame(self.first_page.Tranjestion_btn_frame,fg_color="#D2E3C8")
        button_frame.pack(side="top",pady=10,padx=25,fill="x")

        for text in texts:
            ctk.CTkButton(button_frame,
                          text=text,
                          fg_color="#4F6F52",
                          text_color="white",
                          hover_color="#739072",
                          cursor="hand2",
                          command=lambda t=text:self.Layout_caller(t),
                          ) .pack(pady=10,padx=10,anchor="center",side=side)

    def Layout_caller(self,transaction_name):
        self.first_page.destroy()                 
        if transaction_name == "Check":
            Check_Layout_page_1st(self.window,)
        elif transaction_name == "Loan":
            Loan_BackEnd(self.window)
        else:
            Layout_Transaction_book(self.window,transaction_name)

class First_page(ctk.CTkFrame):
    def __init__(self,window):
        super().__init__(window)
        self.pack(fill="both",expand=True)

        self.window= window
        ctk.CTkLabel(self,text="Money Manager", 
                     font=ctk.CTkFont("Showcard Gothic Regular", 40, "bold")
                     ).place(relx=.15,rely=.04,relwidth=.7)

        self.Tranjestion_btn_frame= ctk.CTkScrollableFrame(self,fg_color="#D2E3C8")
        self.Tranjestion_btn_frame.place(relx=.1,rely=.18,relheight=.48,relwidth=.8)

        self.info_box= ctk.CTkFrame(self,fg_color="#739072")
        self.info_box.place(relx=.173,rely=.7,relheight=.25,relwidth=.66)

        #User_Name
        ctk.CTkLabel(self.info_box,text=self.window.User_Name.get()).pack(anchor="center",pady=6,fill="x",padx=10)

        ctk.CTkLabel(self.info_box,
                     text=f"Date: {self.window.current_date_var.get()}"
                     ).pack(anchor="center",fill="x",pady=3,padx=10)
        self.Ui_mode_on =True
        self.MOde_btn=ctk.CTkButton(self.info_box, text="Trun OFF Ui Mode",fg_color="#4F6F52",hover_color="green", corner_radius=12, cursor="hand2",
                     command=self.change_ui_mode,
                     )
        self.MOde_btn.place(relx=.3,rely=.6, relwidth= .4)
        self.config=configparser.ConfigParser()
        self.config.read(self.window.setting_ini_path)

        
    def change_ui_mode(self,):
        vaid_to_change=False
        if self.Ui_mode_on:
            result = messagebox.askyesno("Confirmation", "Do you want to Enable Command line Mode?")
            if result:
                vaid_to_change=True
                self.Ui_mode_on=False
                self.MOde_btn.configure(text="Trun On Ui Mode")
                self.config.set("Ui_mode","ui_mode","False")
        else:
            result = messagebox.askyesno("Confirmation", "Do you want to Enable Ui Mode Again?")
            if result:
                vaid_to_change=True
                self.Ui_mode_on=True
                self.MOde_btn.configure(text="Trun OFF Ui Mode")
                self.config.set("Ui_mode","ui_mode","True")

        if vaid_to_change:
            with open(self.window.setting_ini_path,'w') as file:
                self.config.write(file)
            pass        




        
        pass    

class Check_Layout_page_1st(ctk.CTkFrame):
    def __init__(self,window,):
        super().__init__(window)
        self.pack(fill="both",expand=True)

        self.window= window

        ctk.CTkLabel(self,text="Money Manager", 
                     font=ctk.CTkFont("Showcard Gothic Regular", 40, "bold")
                     ).place(relx=.15,rely=.03,relwidth=.7)

        # back_btn
        ctk.CTkButton(self,text="↩",cursor="hand2",
                                font=ctk.CTkFont(family="Elephant",size=30),
                                fg_color="Blue",text_color="white",
                                command=self.GetBack).place(relx=.04,rely=.05,relheight=.07,relwidth=.1)

        self.AllElement_frame= ctk.CTkScrollableFrame(self,fg_color="#739072")
        self.AllElement_frame.place(relx=.1,rely=.14,relheight=.3,relwidth=.8)

        self.config=configparser.ConfigParser()
        self.config.read(self.window.setting_ini_path)
        self.elements=(self.config.get("FILE_LOCATION","elements_for_money_track")).split(",")

        ctk.CTkLabel(self.AllElement_frame,text="Current Total Balance On:",
                     font=ctk.CTkFont("Showcard Gothic Regular", 20, "bold")
                     ).pack(pady=5,anchor="center")
        

        for each_element in self.elements:
            solid_trans_name = "_".join(each_element.split(" "))
            amount = float(self.config.get(solid_trans_name, f"{solid_trans_name}_amounts"))

            element_line_frame=ctk.CTkFrame(self.AllElement_frame,height=30)
            element_line_frame.pack(pady=6,padx=5,fill="x")
            ctk.CTkLabel(element_line_frame,text=each_element,
                         corner_radius=0
                     ).place(relx=.02,rely=0,relwidth=.4)
            ctk.CTkLabel(element_line_frame,text=round(amount,4),
            
                     ).place(relx=.44,rely=0,relwidth=.3,)
            
            ctk.CTkButton(element_line_frame,
                          text="See Details",
                          fg_color="#4F6F52",
                          hover_color="#4F6F52",
                          cursor="hand2",
                          command=lambda elem= each_element: self.see_detailes(elem)
                          ).place(relx=.75,rely=0.1,relwidth=.23,relheight=.8)

        element_line_frame=ctk.CTkFrame(self.AllElement_frame,height=30)
        element_line_frame.pack(pady=6,padx=5,fill="x")
        ctk.CTkButton(element_line_frame,text="Click here to See Loan Detailes",
                          fg_color="#4F6F52",
                          hover_color="#31304D",
                          cursor="hand2",
                         corner_radius=10,
                         command=self.loan_label_box_display
                     ).place(relx=.2,rely=0.15,relheight=.7,relwidth=.6)
        self.treeView_strucutre()

    def loan_label_box_display(self):
        self.destroy()
        Loan_check_page_2nd(self.window)
        
    def treeView_strucutre(self):

        self.Tranjestion_btn_frame= ctk.CTkFrame(self,fg_color="#739072")
        self.Tranjestion_btn_frame.place(relx=.1,rely=.47,relheight=.5,relwidth=.8)

        selector_frame=ctk.CTkFrame(self.Tranjestion_btn_frame, fg_color= "transparent")
        selector_frame.place(relx= .05,rely=.05,relheight=.13,relwidth=.9)

        self.ElementOptions=ctk.CTkOptionMenu(selector_frame,values=self.elements,
                                              corner_radius=10,cursor="hand2",
                                              command=self.OPtion_changed)
        self.ElementOptions.place(relx=0,rely=.0,relwidth=.34)

        self.button_30Day=ctk.CTkButton(selector_frame,text="30",cursor="hand2",text_color="black",hover_color='pink',
                                        command=lambda :self.selectdays(30,self.button_30Day))         
        self.button_30Day.place(relx=.38,rely=.0,relwidth=.1,)

        self.button_60Day=ctk.CTkButton(selector_frame,text="60",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.selectdays(60,self.button_60Day))        
        self.button_60Day.place(relx=.5,rely=.0,relwidth=.1,) 

        self.button_90Day=ctk.CTkButton(selector_frame,text="90",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.selectdays(90,self.button_90Day))        
        self.button_90Day.place(relx=.62,rely=.0,relwidth=.1,)
        self.days_button=[self.button_30Day,self.button_60Day,self.button_90Day]

        # Filter
        ctk.CTkButton(selector_frame,text="Filter",
                      cursor="hand2",
                      command=self.filter_Box).place(relx=.75,rely=.0,relwidth=.2,) 

        # Tree view starts
        self.Transaction_info_frame= ctk.CTkScrollableFrame(self.Tranjestion_btn_frame,orientation="horizontal",corner_radius=0)
        self.Transaction_info_frame.place(relx=0,rely=.2,relheight=.8,relwidth=1)
        style = ttk.Style()
        style.configure("Treeview", font=("Times Roman",13), background="#E9EDC9")
        columns = ['SN', 'Date', 'Reason', 'Income', 'Expenses']
        self.treeview = ttk.Treeview(self.Transaction_info_frame, columns=columns, show='headings', selectmode='extended')
        self.treeview.pack(expand=True,fill="both")
        self.treeview.column('SN', width=40)
        self.treeview.column('Date', width=110)
        self.treeview.column('Reason', width=320,anchor=tk.CENTER)
        self.treeview.column('Income', width=90,anchor=tk.CENTER)
        self.treeview.column('Expenses', width=90,anchor=tk.CENTER)

        for col in columns:
            self.treeview.heading(col, text=col)

        heading_font = ('Arial', 15, 'bold')  
        style.configure("Treeview.Heading", background="#e1e1e1", font=heading_font)
        style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])  
        style.configure("Treeview.Treearea", background="#333", foreground="white")
        self.selectdays(30,self.button_30Day)

    def OPtion_changed(self,*args):
        self.selectdays(30,self.button_30Day)

    def see_detailes(self,element):
        self.ElementOptions.set(element)
        self.selectdays(30,self.button_30Day)

    def selectdays(self,day,button_instance):
        for each_btn in self.days_button:
            if each_btn == button_instance:
                    each_btn.configure(fg_color="#A4D0A4",border_color="#DF2E38",border_width=2)
                    self.selected_Days_at_time=day
            else:
                each_btn.configure(border_width=0,fg_color="#1A5D1A")
        self.fill_treeview_with_data(day, None)

    def fill_treeview_with_data(self,day, _Dates_):
        transaction_name=self.ElementOptions.get()
        solid_trans_name = "_".join(transaction_name.split(" "))
        File_path=[]
        directory_path = self.config.get("FILE_LOCATION", "location")
        
        if _Dates_== "random_dates":
            last_month= self.Last_date.strftime("%m")
            all_then_end_date= self.Last_date
        else:
            last_month =(self.config.get(solid_trans_name, f"{solid_trans_name}_last_month"))
            all_then_end_date= (datetime.now()).replace(hour=0, minute=0, second=0, microsecond=0)

        for caller in range(math.ceil(day/30)):
            date_op=(all_then_end_date-timedelta(days=int((caller)*30)))
            front_file_path=f"{directory_path}/{transaction_name}/{transaction_name}'s tranjaction of {date_op.strftime('%Y')}"
            File_path.append(f"{front_file_path}_{date_op.strftime('%m')}.xlsx")
        All_Rows=[]
        for index,each_path in enumerate(File_path):
            try:
                Wb = openpyxl.load_workbook(each_path)
                ws = Wb.active
                RowS= list(ws.rows)
                date_op=(all_then_end_date-timedelta(days=int((index)*30)))
                RowS[0]=(">>","Info of",f"{date_op.strftime('%Y')}, {month_name[int(date_op.strftime('%m'))]}","")
                All_Rows=All_Rows+RowS
            except:pass    

        self.treeview.delete(*self.treeview.get_children())
        for each_cell in All_Rows:
            valued_cell=[]
            if each_cell[1]=="Info of":
                self.treeview.insert("", "end", values=[])  
                valued_cell=each_cell 
            elif str(each_cell[1].value)[0:5] == "Month":
                valued_cell.extend(each_val.value for each_val in each_cell)
                valued_cell.pop(2)

            else:    
                if _Dates_ == "random_dates":
                    d_a_y=(each_cell[1].value).split(",")[0]
                    each_cells_date= datetime(int(time.strftime('%Y')), int(last_month), int(d_a_y), 0,0,0)
                    if self.First_date <= each_cells_date <= self.Last_date:
                        for each_vlaue in each_cell:
                            valued_cell.append(each_vlaue.value)
                        valued_cell.pop(2)
                else:
                    for each_vlaue in each_cell:
                        valued_cell.append(each_vlaue.value)
                    valued_cell.pop(2)
            self.treeview.insert("", "end", values=valued_cell) 


    def filter_Box(self):
        self.filter_frame=ctk.CTkFrame(self,fg_color="#3e3838")
        self.filter_frame.place(relx=.125,rely=.12,relwidth=.75,relheight=.76)
        
        # vars
        present_date = datetime.now()
        self.given_month_ago = present_date - timedelta(days=self.selected_Days_at_time)
        self.calendarBox_displayed=False

        ctk.CTkLabel(self.filter_frame,text=":: Filter ::",text_color="black",
                     fg_color="#8fba9c",
                     corner_radius=10,
                     font=ctk.CTkFont("chiller", 30,"bold")).place(relx=.3,relwidth=.4,rely=.03,)

        ctk.CTkButton(self.filter_frame,text="X",fg_color="red",hover_color="#B80000",
                      command= self.filter_frame.destroy,
                      ).place(relx=.8,rely=.03,relwidth=.18,relheight=.05)
        
        self.ElementOption2=ctk.CTkOptionMenu(self.filter_frame,values=self.elements,
                                              corner_radius=10,cursor="hand2",
                                              command= lambda *args: self.filter_selection_keeper(['no'])
                                              )
        self.ElementOption2.place(relx=.25,rely=.15,relwidth=.5)
        self.Frame_of_search_Function()

        self.Frame_Of_date_Function()
        self.apply_btn=ctk.CTkButton(self.filter_frame,cursor="hand2", text="Apply",fg_color="#7F27FF", hover_color="#9F70FD", command= self.filter_apply_func)
        self.apply_btn.place(relx=.32,rely=.92)

    def date_validate_input(self,char):
        if char == '' :
            return True
        if (char[-1:]).isdigit() or (char[-1:]) == '-' :
            if len(char) <= 10:
                return True
        return False
    
    def Frame_of_search_Function(self):
        outer_frame=ctk.CTkFrame(self.filter_frame, fg_color="grey")
        outer_frame.place(relx=.05,rely=.23,relwidth=.9,relheight= .1)

        self.Search_field= ctk.CTkEntry(outer_frame)
        self.Search_field.place(relx=.02,rely=.05,relheight=.9,relwidth=.8)

        search_btn=ctk.CTkButton(outer_frame,text="🔍", cursor="hand2",fg_color="green",
                                 text_color="blue",
                                 hover_color="light green", font=ctk.CTkFont("chiller", 20,"bold"))
        self.Search_field.place(relx=.02,rely=.05,relheight=.9,relwidth=.8)
        
        pass    

    def Frame_Of_date_Function(self):
        # date_frame
        outer_date_frame= ctk.CTkFrame(self.filter_frame,fg_color="transparent")
        outer_date_frame.place(relx=.035,rely=.35,relwidth=.93,relheight=.55)

        self.scroll_date_frame=ctk.CTkScrollableFrame(outer_date_frame,fg_color="#9ad1aa", corner_radius=0)
        self.scroll_date_frame.pack(expand=True, fill="both")

        date_frame_up_part= ctk.CTkFrame(self.scroll_date_frame,fg_color="#9ad1aa",corner_radius=0,height=85)
        date_frame_up_part.pack(fill= "x")

        ctk.CTkLabel(date_frame_up_part,text="Select Date",font=ctk.CTkFont("Bauhaus 93", 25),text_color="black").place(relx=.32,rely=.02)
        ctk.CTkLabel(date_frame_up_part,text="Start Date:",font=ctk.CTkFont("Times Roman", 20),text_color="black").place(relx=.08,rely=.3,)
        ctk.CTkLabel(date_frame_up_part,text="End Date:"  ,font=ctk.CTkFont("Times Roman", 20),text_color="black").place(relx=.55,rely=.3,)

        start_frame = ctk.CTkFrame(date_frame_up_part,fg_color="#BBE2EC")
        start_frame.place(relx=.08, rely=.66, relwidth=.4, relheight=.26)
        self.start_entry = ctk.CTkEntry(start_frame,corner_radius=0,fg_color="#0A0A0A",border_color="white")
        self.start_entry.place(relx= 0,rely=0, relheight=1, relwidth=.85)
        ctk.CTkButton(start_frame, text="📅", corner_radius=0, 
                      fg_color="#BBE2EC",text_color="blue",hover_color="white", 
                      cursor="hand2",font=('Arial', 20) ,
                      command=lambda :self.show_calendar("S"),
                      ).place(relx= .85,rely=0, relheight=1, relwidth=.15)
        self.start_label= ctk.CTkLabel(date_frame_up_part,text="",fg_color="#1A5D1A")


        end_frame = ctk.CTkFrame(date_frame_up_part,fg_color="#BBE2EC")
        end_frame.place(relx=.55,rely=.66,relwidth=.4,relheight=.26)
        self.end_entry = ctk.CTkEntry(end_frame,corner_radius=0,fg_color="#0A0A0A",border_color="white")
        self.end_entry.place(relx= 0,rely=0, relheight=1, relwidth=.85)
        ctk.CTkButton(end_frame, text="📅", corner_radius=0, 
                      fg_color="#BBE2EC",text_color="blue",hover_color="white", 
                      cursor="hand2",font=('Arial', 20) ,
                      command=lambda :self.show_calendar("E"),
                      ).place(relx= .85,rely=0, relheight=1, relwidth=.15)
        
        self.end_label= ctk.CTkLabel(date_frame_up_part,text="",fg_color="#1A5D1A")
        
        # Validate input and call on_entry_change when the content changes
        validate_cmd = (date_frame_up_part.register(self.date_validate_input),"%P")
        self.start_entry.configure(validate="key", validatecommand=(validate_cmd))
        self.end_entry.configure(validate="key", validatecommand=(validate_cmd))
        self.start_entry.bind("<Return>",self.entry_field_changed)
        self.end_entry.bind("<Return>",self.entry_field_changed)

        self.date_frame_mid_part= ctk.CTkFrame(self.scroll_date_frame,fg_color="transparent",corner_radius=0,)
        self.calendarBox = Calendar(self.date_frame_mid_part, borderwidth=2, font=('Agency fb', 14,"bold"),      
                year=int(self.given_month_ago.strftime("%Y")), month=int(self.given_month_ago.strftime("%m")), day=int(self.given_month_ago.strftime("%d")),                              
                background ="#1A5D1A",
                selectbackground='#00cc00', selectforeground='black',
                firstweekday="sunday",
                showothermonthdays =False,weekenddays=[7,7],
                normalbackground='white', normalforeground='black',
                weekendbackground='white', weekendforeground='black',
                headersbackground='black', headersforeground='white',
                date_pattern='yyyy-mm-dd',
                width=400, height=300
                )
        self.calendarBox.pack(expand=True, fill="both",padx=30)
        ctk.CTkButton(self.date_frame_mid_part,text="X",fg_color="red",hover_color="#B80000",
                      corner_radius=0,
                      command=self.cross_cal,
                      ).place(relx=.94,rely=.03,relwidth=.06,relheight=.1)
        
        # calendar var
        self.First_date=self.given_month_ago.replace(hour=0, minute=0, second=0, microsecond=0)
        self.Last_date=datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) 
        self.start_entry.insert(ctk.END,self.First_date.strftime("%Y-%m-%d"))
        self.end_entry.insert(ctk.END,self.Last_date.strftime("%Y-%m-%d"))


        self.date_frame_down_part= ctk.CTkFrame(self.scroll_date_frame,fg_color="#9ad1aa",corner_radius=0,height=40)
        self.date_frame_down_part.pack(expand=True,fill= "x")

        b30=ctk.CTkButton(self.date_frame_down_part,text="30",cursor="hand2",fg_color="#1A5D1A",
                          text_color="black",hover_color='pink',
                        command=lambda :self.filter_selection_keeper([30,b30])
                                        )
        b30.place(relx=.3,rely=.1,relwidth=.1,)
        b60=ctk.CTkButton(self.date_frame_down_part,text="60",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.filter_selection_keeper([60,b60]) 
                                        )
        b60.place(relx=.45,rely=.1,relwidth=.1,) 
        b90=ctk.CTkButton(self.date_frame_down_part,text="90",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.filter_selection_keeper([90,b90])
                                        )
        b90.place(relx=.6,rely=.1,relwidth=.1,)
        self.btn_list2=[b30,b60,b90]
        self.filter_selection_keeper([self.selected_Days_at_time,self.btn_list2[int(self.selected_Days_at_time/30-1)]])

    def entry_field_changed(self, *args):
        valid_to_change=False
        First_date_= self.start_entry.get().split("-")
        End_date_= self.end_entry.get().split("-")
        try:
            F_year,F_month,F_day=First_date_[0],First_date_[1],First_date_[2]
            E_year,E_month,E_day=End_date_[0],End_date_[1],End_date_[2]
            First_date_= datetime(year=int(F_year), month=int(F_month), day=int(F_day))
            End_date_= datetime(year=int(E_year), month=int(E_month), day=int(E_day))
            valid_to_change= True
        except ValueError as error:
            valid_to_change= False
            if str(error) == "invalid literal for int() with base 10: ''":
                error= "The Structure is Not Accepted as Date format"
                valid_to_change= "Change"

            Notification(self.window,error)  
        except  IndexError as error:
            valid_to_change="Change"
            Notification(self.window, "Plz make sure to full fill your Date format {Y-M-D}")  

        if valid_to_change == "Change":
            self.start_entry.delete(0, ctk.END) 
            self.end_entry.delete(0, ctk.END) 
            self.start_entry.insert(ctk.END,self.First_date.strftime("%Y-%m-%d"))
            self.end_entry.insert(ctk.END,self.Last_date.strftime("%Y-%m-%d"))
        elif valid_to_change == True:
            self.First_date= First_date_
            self.Last_date= End_date_
            try:
                rnads=self.start_or_end
                self.reStructure_cal_box()
            except:
                pass    

    def cross_cal(self):
        self.date_frame_mid_part.pack_forget()
        self.start_entry.configure(fg_color="#0A0A0A",border_color="white")
        self.start_label.place_forget()
        self.end_entry.configure(fg_color="#0A0A0A",border_color="white")
        self.end_label.place_forget()

    def show_calendar(self,start_OR_end):
        self.start_or_end=start_OR_end
        try:
            if last_called_from == start_OR_end:
                if self.calendarBox_displayed:
                    self.calendarBox_displayed=False
                else:
                    self.calendarBox_displayed=True
            else:
                last_called_from=start_OR_end
        except UnboundLocalError:
            self.calendarBox_displayed=True
            last_called_from=start_OR_end

        self.calender_design_look()

        self.reStructure_cal_box()
        self.date_frame_down_part.pack_forget()
        self.date_frame_mid_part.pack(fill="both", expand=True)
        self.date_frame_down_part.pack(expand=True,fill= "x")
        self.calendarBox.configure(tooltipdelay =0,tooltipalpha=0)
        self.calendarBox.bind("<<CalendarSelected>>", self.new_date_selected)

        self.window.after(20, lambda: self.scroll_date_frame._parent_canvas.yview("scroll",-100, "units"))
        self.window.after(20, lambda: self.scroll_date_frame._parent_canvas.yview("scroll",66, "units"))

    def calender_design_look(self):
        if self.start_or_end =="S":
            self.calendarBox.calevent_remove("all")
            self.start_entry.configure(fg_color="#1A5D1A",border_color="#167016")
            self.start_label.place(relx=.08, rely=.92,relwidth=.4, relheight=.08)
            self.end_entry.configure(fg_color="#0A0A0A",border_color="white")
            self.end_label.place_forget()
        else:
            self.end_entry.configure(fg_color="#1A5D1A",border_color="#167016")
            self.end_label.place(relx=.55, rely=.92,relwidth=.38, relheight=.08)    
            self.start_entry.configure(fg_color="#0A0A0A",border_color="white")
            self.start_label.place_forget()

    def reStructure_cal_box(self):
        start,end=self.First_date, self.Last_date
        self.calendarBox.calevent_remove("all")
        self.calendarBox.tag_config("custom_tag", foreground="black", background="#E9EDC9")
        self.calendarBox.tag_config("selected_date_tag", foreground="black", background="#00cc00")
        # from start being called 
        if end != None:
            self.start_entry.delete(0, ctk.END) 
            self.end_entry.delete(0, ctk.END) 
            self.start_entry.insert(ctk.END,self.First_date.strftime("%Y-%m-%d"))
            self.end_entry.insert(ctk.END,self.Last_date.strftime("%Y-%m-%d"))

            while start <= end:
                event_id = self.calendarBox.calevent_create(start,text="", tags=["event_tag","custom_tag"], )
                if (start == end):
                    self.calendarBox.calevent_configure(event_id, tags=["event_tag","selected_date_tag"],text="Selected End date")
                if event_id == 0:
                    self.calendarBox.calevent_configure(event_id, tags=["event_tag","selected_date_tag"],text="Selected Start date")
                start += timedelta(days=1)

            if self.start_or_end =="S":
                self.calendarBox.selection_set(self.First_date.strftime("%Y-%m-%d"))
            else:
                self.calendarBox.selection_set(self.Last_date.strftime("%Y-%m-%d"))

    def new_date_selected(self,event):
        for each_btn in self.btn_list2:
            each_btn.configure(border_width=0,fg_color="#1A5D1A")
        self.No_spefic_date=True    
        if self.start_or_end =="S":
            self.apply_btn.configure(state="disabled")
            self.First_date=datetime.combine((self.calendarBox.selection_get()), datetime.min.time())
            self.calendarBox.calevent_remove("all")
            self.Last_date=None
            self.start_or_end= 'E'
            self.calender_design_look()
            self.calendarBox.selection_set(self.First_date.strftime("%Y-%m-%d"))
            self.start_entry.delete(0, ctk.END) 
            self.end_entry.delete(0, ctk.END) 
            self.start_entry.insert(ctk.END,self.First_date.strftime("%Y-%m-%d"))
            self.end_entry.insert(ctk.END,"0000-00-00")
        else:
            self.apply_btn.configure(state="enabled",cursor="hand2",)
            self.Last_date=datetime.combine((self.calendarBox.selection_get()), datetime.min.time())
            if self.First_date<self.Last_date:
                self.calendarBox.selection_set(self.Last_date.strftime("%Y-%m-%d"))
                self.show_calendar('S')
                self.end_entry.delete(0, ctk.END) 
                self.end_entry.insert(ctk.END,self.Last_date.strftime("%Y-%m-%d"))
            else:
                self.start_or_end='S'
                self.new_date_selected("")

    def filter_selection_keeper(self,days0Relement):   
        self.No_spefic_date=False 
        if type(days0Relement[0]) is int:
            for each_btn in self.btn_list2:
                if each_btn == days0Relement[1]:
                    each_btn.configure(fg_color="#A4D0A4",border_color="#DF2E38",border_width=2)
                    self.select_Days_PARAMERTERS=[days0Relement[0],self.days_button[int(days0Relement[0]/30)-1]]
                else:
                    each_btn.configure(border_width=0,fg_color="#1A5D1A")
            day_total= days0Relement[0]       
        else:
            # self.selectdays(30,self.button_30Day)
            day_total=30
        if self.calendarBox_displayed:
            self.First_date =  (datetime.now()- timedelta(days=day_total)).replace(hour=0, minute=0, second=0, microsecond=0)
            self.Last_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) 
            self.start_or_end="S"
            self.calender_design_look()
            self.reStructure_cal_box()

    def filter_apply_func(self):
        self.filter_frame.destroy()
        self.ElementOptions.set(self.ElementOption2.get())
        if self.No_spefic_date:
            for each_btn in self.days_button:
                each_btn.configure(border_width=0,fg_color="#1A5D1A")
            self.fill_treeview_with_data(int((self.Last_date-self.First_date).days), "random_dates")   
        else:
            self.selectdays(self.select_Days_PARAMERTERS[0],self.select_Days_PARAMERTERS[1])
   
    def GetBack(self):
        self.destroy()
        WinBackEnd(self.window)
        pass

class Loan_check_page_2nd(ctk.CTkFrame):
    def __init__(self,window,):
        super().__init__(window)
        self.pack(fill="both",expand=True)
        self.window= window

        ctk.CTkLabel(self,text="Money Manager", 
                     font=ctk.CTkFont("Showcard Gothic Regular", 40, "bold")
                     ).place(relx=.15,rely=.03,relwidth=.7)

        # back_btn
        ctk.CTkButton(self,text="↩",cursor="hand2",
                                font=ctk.CTkFont(family="Elephant",size=30),
                                fg_color="Blue",text_color="white",
                                command=self.GetBack).place(relx=.04,rely=.05,relheight=.07,relwidth=.1)
        # Variables
        self.Drop_down_options=['All Transaction','Loan Given', 'Loan Recived', 'Not Paid OverDue',"Paid Back Trans"]


        self.upper_section()
        self.down_section()

    def upper_section(self):
        self.AllTangent_list_frame= ctk.CTkScrollableFrame(self,fg_color="#739072")
        self.AllTangent_list_frame.place(relx=.05,rely=.14,relheight=.3,relwidth=.9)

        self.config=configparser.ConfigParser()
        self.config.read(self.window.setting_ini_path)
        self.all_loan_tangents_list= (self.config.get("Loan", "tangents")).split(",")

        if self.all_loan_tangents_list[0]!='':
            self.No_trans_yet=False
            self.all_loan_tangents_list= (self.config.get("Loan", "tangents")).split(",")
            element_line_frame=ctk.CTkFrame(self.AllTangent_list_frame,height=27,corner_radius=0)
            element_line_frame.pack(fill="x")
            ctk.CTkLabel(element_line_frame,text="Tangent Names",
                            corner_radius=0
                        ).place(relx=.02,rely=0,relwidth=.37)
            ctk.CTkLabel(element_line_frame,text="Amt To Give ",
                        ).place(relx=.404,rely=0,relwidth=.2,)

            ctk.CTkLabel(element_line_frame,text="Amt To Take",
                        ).place(relx=.608,rely=0,relwidth=.2,)
            self.detailesBtn=[]
            index=0
            for each_tangent in self.all_loan_tangents_list:
                solid_tangent_name = "_".join(each_tangent.split(" "))
                try:
                    amount = float(self.config.get("Loan", f"{solid_tangent_name}_amt"))                       
                    if amount >= 0:
                        take_amt= amount
                        give_amt=0
                    elif amount< 0:
                        take_amt= 0
                        give_amt=amount
                    element_line_frame=ctk.CTkFrame(self.AllTangent_list_frame,height=30)
                    element_line_frame.pack(pady=6,padx=5,fill="x")
                    ctk.CTkLabel(element_line_frame,text=each_tangent,fg_color="#555555",
                                corner_radius=4
                            ).place(relx=0.005,rely=.04,relwidth=.3938)
                    ctk.CTkLabel(element_line_frame,text=abs(give_amt),fg_color="#555555",
                                corner_radius=3
                            ).place(relx=.404,rely=.04,relwidth=.2,)
                    ctk.CTkLabel(element_line_frame,text=abs(take_amt),fg_color="#555555",
                                corner_radius=3
                            ).place(relx=.608,rely=.04,relwidth=.2,)
                    see_btn=ctk.CTkButton(element_line_frame,
                                text="See Details",
                                fg_color="#4F6F52",
                                hover_color="#4F6F52",
                                cursor="hand2",
                                border_color="red",
                                border_width=0,
                                command=lambda elem= [solid_tangent_name,index]: self.see_detailes(elem)
                                )
                    see_btn.place(relx=.81,rely=0.1,relwidth=.18,relheight=.8) 
                    self.detailesBtn.append(see_btn) 
                    index+=1   
                except: pass   

            self.detailesBtn[0].configure(fg_color="#F8E559",border_width=2,text_color="black",hover_color="sky blue",)
            self.selected_tangent=self.all_loan_tangents_list[0]
        else:
            self.No_trans_yet=True
            Notification(self.window, "Plz First Add Loan Transaction with Tangent","#9A031E",7)    
    
    def see_detailes(self,tangent_info):
        for index,each_btn in enumerate(self.detailesBtn):
            if index == tangent_info[1]:
                each_btn.configure(fg_color="#F8E559",
                                    border_width=2,
                                    text_color="black",
                                    hover_color="sky blue",
                                   )
            else:
                each_btn.configure(fg_color="#4F6F52",border_width=0,hover_color="#4F6F52",text_color="white")    
        self.selected_tangent= tangent_info[0]
        self.D_Doptions.set(self.Drop_down_options[0])
        self.selectdays(360, self.button_360Day)

    def down_section(self):
        self.Tranjestion_btn_frame= ctk.CTkFrame(self,fg_color="#739072",)
        self.Tranjestion_btn_frame.place(relx=.05,rely=.47,relheight=.5,relwidth=.9)
        # Selector
        selector_frame=ctk.CTkFrame(self.Tranjestion_btn_frame,
                                     fg_color= "transparent",)
        selector_frame.place(relx= .01,rely=.02,relheight=.11,relwidth=.98)
        self.D_Doptions=ctk.CTkOptionMenu(selector_frame,values=self.Drop_down_options,
                                              corner_radius=10,cursor="hand2",
                                              command=self.OPtion_changed)
        self.D_Doptions.place(relx=0,rely=.0,relwidth=.325)

        self.button_60Day=ctk.CTkButton(selector_frame,text="60",cursor="hand2",text_color="black",hover_color='pink',fg_color="#1A5D1A",
                                        command=lambda :self.selectdays(60,self.button_60Day))         
        self.button_60Day.place(relx=.34,rely=.0,relwidth=.1,)

        self.button_90Day=ctk.CTkButton(selector_frame,text="90",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.selectdays(90,self.button_90Day))        
        self.button_90Day.place(relx=.46,rely=.0,relwidth=.1,) 

        self.button_120Day=ctk.CTkButton(selector_frame,text="120",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.selectdays(120,self.button_120Day))        
        self.button_120Day.place(relx=.58,rely=.0,relwidth=.1,)

        self.button_360Day=ctk.CTkButton(selector_frame,text="360",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.selectdays(360,self.button_360Day))        
        self.button_360Day.place(relx=.7,rely=.0,relwidth=.1,)
        self.days_button=[self.button_60Day,self.button_90Day,self.button_120Day, self.button_360Day]

        # Filter
        ctk.CTkButton(selector_frame,text="Filter",
                      cursor="hand2",
                      command=self.filter_Box).place(relx=.815,rely=.0,relwidth=.19,) 
        
        self.tree_view()

    def tree_view(self):
        # Tree view starts
        self.Transaction_info_frame= ctk.CTkScrollableFrame(self.Tranjestion_btn_frame,orientation="horizontal",corner_radius=0,)
        self.Transaction_info_frame.place(relx=0,rely=.15,relheight=.85,relwidth=1)
        style = ttk.Style()
        style.configure("Treeview", font=("Times Roman",13), background="#E9EDC9")
        columns = ["SN", "Date", "Due Date", "Reason", "Loan Given", "Loan Taken", "Deposit"]
        self.treeview = ttk.Treeview(self.Transaction_info_frame, columns=columns, show='headings', selectmode='extended')
        self.treeview.pack(expand=True,fill="both")
        self.treeview.column('SN', width=40)
        self.treeview.column('Date', width=120,anchor=tk.CENTER)
        self.treeview.column('Due Date', width=115)
        self.treeview.column('Reason', width=320,anchor=tk.CENTER)
        self.treeview.column('Loan Given', width=120,anchor=tk.CENTER)
        self.treeview.column('Loan Taken', width=150,anchor=tk.CENTER)
        self.treeview.column('Deposit', width=90,anchor=tk.CENTER)

        for col in columns:
            self.treeview.heading(col, text=col)

        heading_font = ('Arial', 15, 'bold')  
        style.configure("Treeview.Heading", background="#e1e1e1", font=heading_font)
        style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])  
        style.configure("Treeview.Treearea", background="#333", foreground="white")

        self.selectdays(360, self.button_360Day)
        pass
    
    def OPtion_changed(self,*args):
        if self.D_Doptions.get() =="All Transaction":
            self.selectdays(360,self.button_360Day)
        else:
            self.selectdays(60,self.button_60Day)

    def selectdays(self,day,button_instance):
        for each_btn in self.days_button:
            if each_btn == button_instance:
                    each_btn.configure(fg_color="#A4D0A4",border_color="#DF2E38",border_width=2)
                    self.selected_Days_at_time=day
            else:
                    each_btn.configure(border_width=0,fg_color="#1A5D1A")

        self.fill_treeview_with_data(day, None)

    def fill_treeview_with_data(self,day, _Dates_):
        if not self.No_trans_yet:
            directory_path = self.config.get("FILE_LOCATION", "location")
            File_path=f"{directory_path}/Loan/Loan_with_{self.selected_tangent}.xlsx"
            
            if _Dates_== "random_dates":
                last_month= self.Last_date.strftime("%m")
                all_then_end_date= self.Last_date
                all_then_start_date= self.First_date

            else:
                last_month =datetime.now().strftime("%m")
                all_then_end_date= (datetime.now()).replace(hour=0, minute=0, second=0, microsecond=0)
                all_then_start_date=(all_then_end_date-timedelta(days=int(((math.ceil(day/30))+1)*30)))        
    
            # Each row of particular tangent is being called and being saved under All_Rows   
            try:
                Wb = openpyxl.load_workbook(File_path)
                ws = Wb.active
                AllCell= list(ws.rows)
                AllCell.pop(0)
                AllCell=self.Check_DropDown_option_check(AllCell)
                Running_transOn_date= 0
                self.treeview.delete(*self.treeview.get_children())
                for each_row in AllCell:
                    valued_cell=[]  
                    valid_to_insert=False
                    this_rowsDate=  (each_row[1]).value[:-7]

                    rows_date= ((each_row[1]).value[:-4]).split("-")
                    rows_with_date_instance= datetime(int(rows_date[0]),int(rows_date[1]),int(rows_date[2]))
                    if all_then_start_date <=rows_with_date_instance<=all_then_end_date :
                            for each_vlaue in each_row:
                                valid_to_insert=True
                                valued_cell.append(each_vlaue.value)
                    if this_rowsDate != Running_transOn_date and valid_to_insert:
                        Running_transOn_date= this_rowsDate
                        new_values=[">>","-----------","Info of",f"{this_rowsDate[:-3]}, {month_name[int(this_rowsDate[5:])]}","---","---"]
                        self.treeview.insert("", "end", values=[])  
                        self.treeview.insert("", "end", values=new_values)             
                    if valued_cell:    
                        self.treeview.insert("", "end", values=valued_cell) 
            except:pass

    def Check_DropDown_option_check(self,AllCell):
        dd_value=self.D_Doptions.get()
        if dd_value =="All Transaction":
            return AllCell
        elif dd_value =="Loan Given":
            new_cell=[]
            for each_row in AllCell:
                if each_row[4].value !=0:
                    new_cell.append(each_row)
            return new_cell        
        elif dd_value =="Loan Recived":
            new_cell=[]
            for each_row in AllCell:
                if each_row[5].value !=0:
                    new_cell.append(each_row)
            return new_cell              
        elif dd_value =="Not Paid OverDue":
            new_cell=[]
            for each_row in AllCell:
                STRdate= ((each_row[1]).value[:-4]).split("-")
                due_date= datetime(year=int(STRdate[0]),month=int(STRdate[1]), day=int(STRdate[2]))
                presentdate= datetime.now()
                if presentdate < due_date:
                    new_cell.append(each_row)
            return new_cell 
        elif dd_value =="Paid Back Trans":
            new_cell=[]
            for each_row in AllCell:
                if each_row[6].value:
                    new_cell.append(each_row)
            return new_cell             

    def filter_Box(self):
        self.filter_frame=ctk.CTkFrame(self,fg_color="#3e3838")
        self.filter_frame.place(relx=.125,rely=.12,relwidth=.75,relheight=.76)
        
        # vars
        present_date = datetime.now()
        self.given_month_ago = present_date - timedelta(days=self.selected_Days_at_time)
        self.calendarBox_displayed=False

        ctk.CTkLabel(self.filter_frame,text=":: Filter ::",text_color="black",
                     fg_color="#8fba9c",
                     corner_radius=10,
                     font=ctk.CTkFont("chiller", 30,"bold")).place(relx=.3,relwidth=.4,rely=.03,)

        ctk.CTkButton(self.filter_frame,text="X",fg_color="red",hover_color="#B80000",
                      command= self.filter_frame.destroy,
                      ).place(relx=.8,rely=.03,relwidth=.18,relheight=.05)
        
        self.ElementOption2=ctk.CTkOptionMenu(self.filter_frame,values=self.Drop_down_options,
                                              corner_radius=10,cursor="hand2",
                                              command= lambda *args: self.filter_selection_keeper(['no'])
                                              )
        self.ElementOption2.place(relx=.05,rely=.15,relwidth=.4)

        self.Tangent_choose = ctk.CTkOptionMenu(self.filter_frame,values=self.all_loan_tangents_list,
                                              corner_radius=10,cursor="hand2",
                                              command= lambda *args: self.filter_selection_keeper(['no'])
                                              )
        self.Tangent_choose.place(relx= .55, rely = .15,relwidth = .4)

        self.Frame_of_search_Function()
        self.Frame_Of_date_Function()

        self.apply_btn=ctk.CTkButton(self.filter_frame,cursor="hand2", text="Apply",fg_color="#7F27FF", hover_color="#9F70FD", command= self.filter_apply_func)
        self.apply_btn.place(relx=.32,rely=.92)

    def date_validate_input(self,char):
        if char == '' :
            return True
        if (char[-1:]).isdigit() or (char[-1:]) == '-' :
            if len(char) <= 10:
                return True
        return False
    
    def Frame_of_search_Function(self):
        outer_frame=ctk.CTkFrame(self.filter_frame, fg_color="grey")
        outer_frame.place(relx=.05,rely=.24,relwidth=.9,relheight= .08)

        self.Search_field= ctk.CTkEntry(outer_frame)
        self.Search_field.place(relx=.02,rely=.05,relheight=.9,relwidth=.8)

        search_btn=ctk.CTkButton(outer_frame,text="🔍", cursor="hand2",fg_color="green",
                                 text_color="blue",
                                 hover_color="light green", font=ctk.CTkFont("chiller", 20,"bold"))
        search_btn.place(relx=.85,rely=.05,relheight=.9,relwidth=.12)
        pass    

    def Frame_Of_date_Function(self):
        # date_frame
        outer_date_frame= ctk.CTkFrame(self.filter_frame,fg_color="transparent")
        outer_date_frame.place(relx=.035,rely=.35,relwidth=.93,relheight=.55)

        self.scroll_date_frame=ctk.CTkScrollableFrame(outer_date_frame,fg_color="#9ad1aa", corner_radius=0)
        self.scroll_date_frame.pack(expand=True, fill="both")

        date_frame_up_part= ctk.CTkFrame(self.scroll_date_frame,fg_color="#9ad1aa",corner_radius=0,height=85)
        date_frame_up_part.pack(fill= "x")

        ctk.CTkLabel(date_frame_up_part,text="Select Date",font=ctk.CTkFont("Bauhaus 93", 25),text_color="black").place(relx=.32,rely=.02)
        ctk.CTkLabel(date_frame_up_part,text="Start Date:",font=ctk.CTkFont("Times Roman", 20),text_color="black").place(relx=.08,rely=.3,)
        ctk.CTkLabel(date_frame_up_part,text="End Date:"  ,font=ctk.CTkFont("Times Roman", 20),text_color="black").place(relx=.55,rely=.3,)

        start_frame = ctk.CTkFrame(date_frame_up_part,fg_color="#BBE2EC")
        start_frame.place(relx=.08, rely=.66, relwidth=.4, relheight=.26)
        self.start_entry = ctk.CTkEntry(start_frame,corner_radius=0,fg_color="#0A0A0A",border_color="white")
        self.start_entry.place(relx= 0,rely=0, relheight=1, relwidth=.85)
        ctk.CTkButton(start_frame, text="📅", corner_radius=0, 
                      fg_color="#BBE2EC",text_color="blue",hover_color="white", 
                      cursor="hand2",font=('Arial', 20) ,
                      command=lambda :self.show_calendar("S"),
                      ).place(relx= .85,rely=0, relheight=1, relwidth=.15)
        self.start_label= ctk.CTkLabel(date_frame_up_part,text="",fg_color="#1A5D1A")


        end_frame = ctk.CTkFrame(date_frame_up_part,fg_color="#BBE2EC")
        end_frame.place(relx=.55,rely=.66,relwidth=.4,relheight=.26)
        self.end_entry = ctk.CTkEntry(end_frame,corner_radius=0,fg_color="#0A0A0A",border_color="white")
        self.end_entry.place(relx= 0,rely=0, relheight=1, relwidth=.85)
        ctk.CTkButton(end_frame, text="📅", corner_radius=0, 
                      fg_color="#BBE2EC",text_color="blue",hover_color="white", 
                      cursor="hand2",font=('Arial', 20) ,
                      command=lambda :self.show_calendar("E"),
                      ).place(relx= .85,rely=0, relheight=1, relwidth=.15)
        
        self.end_label= ctk.CTkLabel(date_frame_up_part,text="",fg_color="#1A5D1A")
        
        # Validate input and call on_entry_change when the content changes
        validate_cmd = (date_frame_up_part.register(self.date_validate_input),"%P")
        self.start_entry.configure(validate="key", validatecommand=(validate_cmd))
        self.end_entry.configure(validate="key", validatecommand=(validate_cmd))
        self.start_entry.bind("<Return>",self.entry_field_changed)
        self.end_entry.bind("<Return>",self.entry_field_changed)

        self.date_frame_mid_part= ctk.CTkFrame(self.scroll_date_frame,fg_color="transparent",corner_radius=0,)
        self.calendarBox = Calendar(self.date_frame_mid_part, borderwidth=2, font=('Agency fb', 14,"bold"),      
                year=int(self.given_month_ago.strftime("%Y")), month=int(self.given_month_ago.strftime("%m")), day=int(self.given_month_ago.strftime("%d")),                              
                background ="#1A5D1A",
                selectbackground='#00cc00', selectforeground='black',
                firstweekday="sunday",
                showothermonthdays =False,weekenddays=[7,7],
                normalbackground='white', normalforeground='black',
                weekendbackground='white', weekendforeground='black',
                headersbackground='black', headersforeground='white',
                date_pattern='yyyy-mm-dd',
                width=400, height=300
                )
        self.calendarBox.pack(expand=True, fill="both",padx=30)
        ctk.CTkButton(self.date_frame_mid_part,text="X",fg_color="red",hover_color="#B80000",
                      corner_radius=0,
                      command=self.cross_cal,
                      ).place(relx=.94,rely=.03,relwidth=.06,relheight=.1)
        
        # calendar var
        self.First_date=self.given_month_ago.replace(hour=0, minute=0, second=0, microsecond=0)
        self.Last_date=datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) 
        self.start_entry.insert(ctk.END,self.First_date.strftime("%Y-%m-%d"))
        self.end_entry.insert(ctk.END,self.Last_date.strftime("%Y-%m-%d"))


        self.date_frame_down_part= ctk.CTkFrame(self.scroll_date_frame,fg_color="#9ad1aa",corner_radius=0,height=40)
        self.date_frame_down_part.pack(expand=True,fill= "x")

        b60=ctk.CTkButton(self.date_frame_down_part,text="60",cursor="hand2",fg_color="#1A5D1A",
                          text_color="black",hover_color='pink',
                        command=lambda :self.filter_selection_keeper([60,b60])
                                        )
        b60.place(relx=.2,rely=.1,relwidth=.15,)
        b90=ctk.CTkButton(self.date_frame_down_part,text="90",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.filter_selection_keeper([90,b90]) 
                                        )
        b90.place(relx=.37,rely=.1,relwidth=.15,) 
        b120=ctk.CTkButton(self.date_frame_down_part,text="120",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.filter_selection_keeper([120,b120])
                                        )
        b120.place(relx=.54,rely=.1,relwidth=.15,)

        b360=ctk.CTkButton(self.date_frame_down_part,text="360",cursor="hand2",fg_color="#1A5D1A",text_color="black",hover_color='pink',
                                        command=lambda :self.filter_selection_keeper([360,b360])
                                        )
        b360.place(relx=.71,rely=.1,relwidth=.15,)
        self.btn_list2=[b60,b90,b120, b360]
        self.filter_selection_keeper([self.selected_Days_at_time,self.btn_list2[3 if self.selected_Days_at_time == 360 else int(self.selected_Days_at_time / 30 - 2)
]])

    def entry_field_changed(self, *args):
        valid_to_change=False
        First_date_= self.start_entry.get().split("-")
        End_date_= self.end_entry.get().split("-")
        try:
            F_year,F_month,F_day=First_date_[0],First_date_[1],First_date_[2]
            E_year,E_month,E_day=End_date_[0],End_date_[1],End_date_[2]
            First_date_= datetime(year=int(F_year), month=int(F_month), day=int(F_day))
            End_date_= datetime(year=int(E_year), month=int(E_month), day=int(E_day))
            valid_to_change= True
        except ValueError as error:
            valid_to_change= False
            if str(error) == "invalid literal for int() with base 10: ''":
                error= "The Structure is Not Accepted as Date format"
                valid_to_change= "Change"

            Notification(self.window,error)  
        except  IndexError as error:
            valid_to_change="Change"
            Notification(self.window, "Plz make sure to full fill your Date format {Y-M-D}")  

        if valid_to_change == "Change":
            self.start_entry.delete(0, ctk.END) 
            self.end_entry.delete(0, ctk.END) 
            self.start_entry.insert(ctk.END,self.First_date.strftime("%Y-%m-%d"))
            self.end_entry.insert(ctk.END,self.Last_date.strftime("%Y-%m-%d"))
        elif valid_to_change == True:
            self.First_date= First_date_
            self.Last_date= End_date_
            try:
                rnads=self.start_or_end
                self.reStructure_cal_box()
            except:
                pass    

    def cross_cal(self):
        self.date_frame_mid_part.pack_forget()
        self.start_entry.configure(fg_color="#0A0A0A",border_color="white")
        self.start_label.place_forget()
        self.end_entry.configure(fg_color="#0A0A0A",border_color="white")
        self.end_label.place_forget()

    def show_calendar(self,start_OR_end):
        self.start_or_end=start_OR_end
        try:
            if last_called_from == start_OR_end:
                if self.calendarBox_displayed:
                    self.calendarBox_displayed=False
                else:
                    self.calendarBox_displayed=True
            else:
                last_called_from=start_OR_end
        except UnboundLocalError:
            self.calendarBox_displayed=True
            last_called_from=start_OR_end

        self.calender_design_look()

        self.reStructure_cal_box()
        self.date_frame_down_part.pack_forget()
        self.date_frame_mid_part.pack(fill="both", expand=True)
        self.date_frame_down_part.pack(expand=True,fill= "x")
        self.calendarBox.configure(tooltipdelay =0,tooltipalpha=0)
        self.calendarBox.bind("<<CalendarSelected>>", self.new_date_selected)

        self.window.after(20, lambda: self.scroll_date_frame._parent_canvas.yview("scroll",-100, "units"))
        self.window.after(20, lambda: self.scroll_date_frame._parent_canvas.yview("scroll",66, "units"))

    def calender_design_look(self):
        if self.start_or_end =="S":
            self.calendarBox.calevent_remove("all")
            self.start_entry.configure(fg_color="#1A5D1A",border_color="#167016")
            self.start_label.place(relx=.08, rely=.92,relwidth=.4, relheight=.08)
            self.end_entry.configure(fg_color="#0A0A0A",border_color="white")
            self.end_label.place_forget()
        else:
            self.end_entry.configure(fg_color="#1A5D1A",border_color="#167016")
            self.end_label.place(relx=.55, rely=.92,relwidth=.38, relheight=.08)    
            self.start_entry.configure(fg_color="#0A0A0A",border_color="white")
            self.start_label.place_forget()

    def reStructure_cal_box(self):
        start,end=self.First_date, self.Last_date
        self.calendarBox.calevent_remove("all")
        self.calendarBox.tag_config("custom_tag", foreground="black", background="#E9EDC9")
        self.calendarBox.tag_config("selected_date_tag", foreground="black", background="#00cc00")
        # from start being called 
        if end != None:
            self.start_entry.delete(0, ctk.END) 
            self.end_entry.delete(0, ctk.END) 
            self.start_entry.insert(ctk.END,self.First_date.strftime("%Y-%m-%d"))
            self.end_entry.insert(ctk.END,self.Last_date.strftime("%Y-%m-%d"))

            while start <= end:
                event_id = self.calendarBox.calevent_create(start,text="", tags=["event_tag","custom_tag"], )
                if (start == end):
                    self.calendarBox.calevent_configure(event_id, tags=["event_tag","selected_date_tag"],text="Selected End date")
                if event_id == 0:
                    self.calendarBox.calevent_configure(event_id, tags=["event_tag","selected_date_tag"],text="Selected Start date")
                start += timedelta(days=1)

            if self.start_or_end =="S":
                self.calendarBox.selection_set(self.First_date.strftime("%Y-%m-%d"))
            else:
                self.calendarBox.selection_set(self.Last_date.strftime("%Y-%m-%d"))

    def new_date_selected(self,event):
        for each_btn in self.btn_list2:
            each_btn.configure(border_width=0,fg_color="#1A5D1A")
        self.No_spefic_date=True    
        if self.start_or_end =="S":
            self.apply_btn.configure(state="disabled")
            self.First_date=datetime.combine((self.calendarBox.selection_get()), datetime.min.time())
            self.calendarBox.calevent_remove("all")
            self.Last_date=None
            self.start_or_end= 'E'
            self.calender_design_look()
            self.calendarBox.selection_set(self.First_date.strftime("%Y-%m-%d"))
            self.start_entry.delete(0, ctk.END) 
            self.end_entry.delete(0, ctk.END) 
            self.start_entry.insert(ctk.END,self.First_date.strftime("%Y-%m-%d"))
            self.end_entry.insert(ctk.END,"0000-00-00")
        else:
            self.apply_btn.configure(state="enabled",cursor="hand2",)
            self.Last_date=datetime.combine((self.calendarBox.selection_get()), datetime.min.time())
            if self.First_date<self.Last_date:
                self.calendarBox.selection_set(self.Last_date.strftime("%Y-%m-%d"))
                self.show_calendar('S')
                self.end_entry.delete(0, ctk.END) 
                self.end_entry.insert(ctk.END,self.Last_date.strftime("%Y-%m-%d"))
            else:
                self.start_or_end='S'
                self.new_date_selected("")

    def filter_selection_keeper(self,days0Relement):   
        self.No_spefic_date=False 
        if type(days0Relement[0]) is int:
            for each_btn in self.btn_list2:
                if each_btn == days0Relement[1]:
                    each_btn.configure(fg_color="#A4D0A4",border_color="#DF2E38",border_width=2)
                    self.select_Days_PARAMERTERS=[days0Relement[0],self.days_button[3 if days0Relement[0] == 360 else int(days0Relement[0] / 30 - 2)]]
                else:
                    each_btn.configure(border_width=0,fg_color="#1A5D1A")
            day_total= days0Relement[0]       
        else:
            day_total=60
        if self.calendarBox_displayed:
            self.First_date =  (datetime.now()- timedelta(days=day_total)).replace(hour=0, minute=0, second=0, microsecond=0)
            self.Last_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) 
            self.start_or_end="S"
            self.calender_design_look()
            self.reStructure_cal_box()

    def filter_apply_func(self):
        self.filter_frame.destroy()
        self.D_Doptions.set(self.ElementOption2.get())

        # for tangent name fixing
        tangent= self.Tangent_choose.get()
        solid_tangent_name = "_".join(tangent.split(" "))
        index= self.all_loan_tangents_list.index(tangent)
        self.see_detailes([solid_tangent_name,index])

        if self.No_spefic_date:
            for each_btn in self.days_button:
                each_btn.configure(border_width=0,fg_color="#1A5D1A")
            self.fill_treeview_with_data(int((self.Last_date-self.First_date).days), "random_dates")   
        else:
            self.selectdays(self.select_Days_PARAMERTERS[0],self.select_Days_PARAMERTERS[1])
   
    def GetBack(self):
        self.destroy()
        Check_Layout_page_1st(self.window)

class Layout_Transaction_book():
    def __init__(self, window, transaction_name):
        self.window=window
        self.transaction_name=transaction_name

        self.Main_frame=ctk.CTkFrame(self.window,fg_color="transparent")
        self.Main_frame.pack(expand=True,fill="both")

        ctk.CTkLabel(self.Main_frame,text="Money Manager",
                     fg_color="transparent",
                     font=ctk.CTkFont("Showcard Gothic Regular", 40, "bold")
                     ).place(relx=.15,rely=.03,relwidth=.7,relheight=.1)
        # back_btn
        ctk.CTkButton(self.Main_frame,text="↩",cursor="hand2",
                                font=ctk.CTkFont(family="Elephant",size=30),
                                fg_color="Blue",text_color="white",
                                command=self.GetBack
                                ).place(relx=.04,rely=.05,relheight=.07,relwidth=.1)
        
        Menu_Frame=ctk.CTkTabview(self.Main_frame,
                                  segmented_button_selected_color="green",
                                  segmented_button_selected_hover_color="#86A789",
                                  segmented_button_unselected_color="#D2E3C8",
                                  text_color="black")
        Menu_Frame.place(relx=.05,relwidth=.9,rely=.12,relheight=.85)

        Menu_Frame.add("Inflow")
        Menu_Frame.add("Outflow")
        self.degine_frame1=ctk.CTkFrame(Menu_Frame.tab("Inflow"),fg_color="transparent")
        self.degine_frame1.pack(expand=True,fill="both")
        self.degine_frame2=ctk.CTkFrame(Menu_Frame.tab("Outflow"),fg_color="transparent")
        self.degine_frame2.pack(expand=True,fill="both")

        Layout_box(self,self.degine_frame1,"ADD")
        Layout_box(self,self.degine_frame2,"SUB")

    def GetBack(self):
        self.Main_frame.destroy()
        Backend_First_page(self.window)

class Layout_box():
    def __init__(self,parent,Design_frame,add_or_sub) -> None:
        self.window=parent.window
        self.transaction_name=parent.transaction_name
        self.transaction_full_details=[]
        self.selected_keyword_list=[]
        self.transaction_box_value=[]
        self.total_transation_frame=[]
        self.keywordButtonlist=[]
        self.reason_added=0
        self.add_or_sub=add_or_sub
        if self.add_or_sub =="ADD":
            self.src_res="Source  :"
            self.conner_text="In Rs, [ Plus ➕ ]"
        else:
            self.src_res="Reason  :"
            self.conner_text="In Rs, [ Minus ➖ ]"

        self.config=configparser.ConfigParser()
        self.config.read(self.window.setting_ini_path)
        self.setting_keyword_name=(self.config.get("Keywords","Keywords")).split(",")

        self.Upper_Half(Design_frame)
        self.Lower_Half(Design_frame)

    def validate_amount(self, event):
        if event.char in ['', '\b',"."]:return
        if not event.char.isdigit():
            self.amount_entry_box.bell()
            return 'break'

    def handle(self, *args):
        pass        

    def EnteR(self,*args):
        self.add_transation()
        pass

    def on_focus_in(self, *args):
        if self.amount_entry_box.get()==self.placeholder:
            self.amount_entry_box.delete(0, "end")      

    def Upper_Half(self,Design_frame):
        ctk.CTkLabel(Design_frame,text=f"{self.transaction_name}'s Transation",
                     font=ctk.CTkFont("Agency fb", 30, "bold"),
                     text_color="red",anchor="center"
                     ).place(relx=0,rely=0,relheight=.1,relwidth=1)

        upper_half_frame=ctk.CTkFrame(Design_frame,corner_radius=0,)                             
        upper_half_frame.place(relx=0,rely=.1,relheight=.5,relwidth=1)

        # input frame
        input_frame=ctk.CTkFrame(upper_half_frame,corner_radius=0)
        input_frame.place(relx=0,rely=.0,relwidth=1,relheight=.36)

        #source
        ctk.CTkLabel(input_frame,text=self.src_res).place(relx=.12,rely=.1)
        self.reason_entry_box=ctk.CTkEntry(input_frame,)
        self.reason_entry_box.place(relx=.25,rely=.1,relwidth=.6)

        #amount
        ctk.CTkLabel(input_frame,text="Amount :",).place(relx=.12,rely=.55)
        self.placeholder = self.conner_text
        self.amount_entry_box=ctk.CTkEntry(input_frame,)
        self.amount_entry_box.place(relx=.25,rely=.55,relwidth=.6)
        self.amount_entry_box.insert(0,self.placeholder)
        self.amount_entry_box.bind("<FocusIn>",self.on_focus_in)

        self.amount_entry_box.bind('<Key>', self.validate_amount)
        self.amount_entry_box.bind('<Tab>', self.handle)
        self.amount_entry_box.bind('<Control-v>', self.handle)
        self.amount_entry_box.bind('<Control-a>', self.handle)
        self.amount_entry_box.bind('<Control-c>', self.handle)
        self.amount_entry_box.bind('<Return>', self.EnteR)

        #Amount_resourece add button
        ctk.CTkButton(input_frame,text="\u2795 ",
                              fg_color="green",
                              hover_color="red",
                              cursor="hand2",
                              command=self.add_transation,
                              corner_radius=0
                              ).place(relx=.87,rely=.25,relheight=.5,relwidth=.1)

        #key word frame
        keyword_frame=ctk.CTkFrame(upper_half_frame,corner_radius=0,)
        keyword_frame.place(relx=0,rely=.36,relwidth=1,relheight=.64)

        #key Word adder
        keyword_adder_frame=ctk.CTkFrame(keyword_frame,corner_radius=0)
        keyword_adder_frame.place(relx=.0,rely=.05,relwidth=1,relheight=.25)

        ctk.CTkLabel(keyword_adder_frame,text="Keywords",
                     text_color="White",
                     fg_color="#2B2B2B",
                     corner_radius=0,
                     font=ctk.CTkFont("Algerian", 20, "bold")
                     ).place(relx=.02,rely=0,relwidth=.3)

        self.entrybox_keyword_adder=ctk.CTkEntry(keyword_adder_frame,)
        self.entrybox_keyword_adder.place(relx=.386,rely=0,relwidth=.5)

        self.entrybox_keyword_adder.bind("<Return>",self.keyword_adder)

        ctk.CTkButton(keyword_adder_frame,text="+",fg_color="#864AF9",
                      command=self.keyword_adder
                      ).place(relx=.9,rely=0,relwidth=.07)

        #list of keys
        self.list_of_keywords_frame=ctk.CTkScrollableFrame(keyword_frame, 
                                                      orientation="horizontal",
                                                      corner_radius=0)
        self.list_of_keywords_frame.place(relx=0,rely=.25,relwidth=1,relheight=.75)

        self.list_1st_part=ctk.CTkFrame(self.list_of_keywords_frame,)
        self.list_1st_part.pack(pady=8,padx=30)
        self.list_2nd_part=ctk.CTkFrame(self.list_of_keywords_frame,)
        self.list_2nd_part.pack(pady=2,side="left",padx=30)

        #keyword deleter
        ctk.CTkButton(self.list_of_keywords_frame,text="\U0001F5D1",
                      cursor="hand2",
                      hover_color="#FF004D",
                      command=self.remove_keyword_btn,
                      width=20,
                      fg_color="red").place(x=5,y=28)

        self.keyword_adder(Add_new=False)
        
    def keyword_adder(self,Add_new=True):        
        if Add_new:
            new_keyword=self.entrybox_keyword_adder.get()
            if new_keyword=='':
                Notification(self.window,"Write some text here First")
                return
            elif "," in new_keyword:
                Notification(self.window,"PLz Dont keep \",\" in your Entry  field")
                return
    
            else:    
                for child in (self.list_1st_part.winfo_children()):
                    child.destroy()
                for child in (self.list_2nd_part.winfo_children()):
                    child.destroy() 
                self.keywordButtonlist=[]
                self.setting_keyword_name.append(new_keyword)
                self.entrybox_keyword_adder.delete(0,'end')
                self.keywordsStructure_caller()    
                self.selected_keyword_list=[]
        else:
            self.keywordsStructure_caller()      

    def keywordsStructure_caller(self):
        for i in range(0,len(self.setting_keyword_name),2):
            if i== len(self.setting_keyword_name)-1:
                self.keywords_structure(self.list_1st_part,self.setting_keyword_name[i],i)
            else:    
                self.keywords_structure(self.list_1st_part,self.setting_keyword_name[i],i)
                self.keywords_structure(self.list_2nd_part,self.setting_keyword_name[i+1],i+1)

    def keywords_structure(self,parent,text,index):
            keyword_button=ctk.CTkButton(parent,
                          text=text,
                          fg_color="#4F6F52",
                          text_color="white",
                          hover_color="#739072",
                          cursor="hand2",
                          height=25,
                          command=lambda :self.select_keyword(index)
                          )            
            keyword_button.pack(padx=10,side="left")  
            self.keywordButtonlist.append(keyword_button)

    def select_keyword(self,index):
        key_word=self.keywordButtonlist[index]
        if key_word in self.selected_keyword_list:
            key_word.configure(border_width=0,fg_color="#4F6F52")
            self.selected_keyword_list.remove(key_word)
        else:
            self.selected_keyword_list.append(key_word)
            key_word.configure(border_width=2,border_color="yellow",fg_color="GREEN")

    def remove_keyword_btn(self, normal_call=True):
        if self.selected_keyword_list:
            for each_btn in self.selected_keyword_list:
                self.setting_keyword_name.remove(each_btn.cget("text"))
                self.selected_keyword_list.remove(each_btn)
                each_btn.destroy()
            self.remove_keyword_btn(normal_call=False)
        else:
            if normal_call:
                Notification(self.window,'Plz select keywords and hit Delete')                        
    
    def Lower_Half(self,Design_frame):
        lower_half=ctk.CTkFrame(Design_frame,corner_radius=0,)                             
        lower_half.place(relx=0,rely=.6,relheight=.4,relwidth=1)

        self.list0fTransaction=ctk.CTkScrollableFrame(lower_half,corner_radius=0,)
        self.list0fTransaction.place(relx=0,rely=.05,relwidth=1,relheight=.82)

        title_frame=ctk.CTkFrame(self.list0fTransaction,height=20,corner_radius=0)
        title_frame.pack(side="top",fill="x")

        ctk.CTkLabel(title_frame,text="N0.").place(relx=0 ,rely=0,relwidth=.1,relheight=.6)
        ctk.CTkLabel(title_frame,text="Reason Here",).place(relx=.1,rely=0,relwidth=.75,relheight=.6)
        ctk.CTkLabel(title_frame,text="Amount").place(relx=.78,rely=0,relwidth=.12,relheight=.6)

        button_frame=ctk.CTkFrame(lower_half,corner_radius=0,)
        button_frame.place(relx=0,rely=.83,relwidth=1,relheight=.18)

        send_button  =ctk.CTkButton(button_frame,text="Send",fg_color="#4F6F52",
                                    command=self.Send_to_calculator,
                                    hover_color="#19A7CE",cursor="hand2")
        send_button  .place(relx=.25,rely=.0,relwidth=.2)
        Revert_button=ctk.CTkButton(button_frame,
                                    text="Revert",
                                    fg_color="#4F6F52",
                                    hover_color="#ED2B2A",
                                    command= self.revert_all_transation_list,
                                    cursor="hand2")
        Revert_button.place(relx=.55,rely=.0,relwidth=.2)       
        pass    

    def add_transation(self):
        valid_input = True
        if not self.reason_entry_box.get():
            Notification(self.window,f"Add your {self.src_res} First !!!")
            valid_input = False
        elif not self.amount_entry_box.get() or self.amount_entry_box.get() == self.placeholder:
            Notification(self.window,"Add your amount First !!!")   
            valid_input = False

        elif len(self.selected_keyword_list) <2:
            Notification(self.window,"Select atleast 2 Key word !!!")
            valid_input = False

        else:
            try:
                float(self.amount_entry_box.get())
            except:    
                Notification(self.window,f"In amount section String found, remove it !!!")
                valid_input = False

        if valid_input:
            self.reason_added+=1
            reason_source=self.reason_entry_box.get()
            amount=self.amount_entry_box.get()
            transaction_frame=ctk.CTkFrame(self.list0fTransaction,height=40)
            transaction_frame.pack(side="top",fill="x",pady=5,padx=5)
            ctk.CTkLabel(transaction_frame,text=self.reason_added).place(relx=0 ,rely=0.2,relwidth=.05,relheight=.6)
            ctk.CTkLabel(transaction_frame,text=reason_source).place(relx=.05,rely=0.2,relwidth=.72,relheight=.6)
            ctk.CTkLabel(transaction_frame,text=amount,fg_color="black").place(relx=.78,rely=0.2,relwidth=.14,relheight=.6)
            ctk.CTkButton(transaction_frame,
                          text="\u2796 ",
                          fg_color="red",
                          cursor="hand2",
                          hover_color="#EC255A",
                          corner_radius=0,
                          font=("Arial", 18,"bold"),
                          command=lambda index=self.reason_added,f=transaction_frame:self.delete_transation_box(f)
                          ).place(relx=.93,rely=0.2,relwidth=.06,relheight=.6)   
            
            self.amount_entry_box.delete(0,"end")
            self.total_transation_frame.append(transaction_frame)
            self.keyword_collection=[]
            for each_keywords in self.selected_keyword_list:
                self.keyword_collection.append(each_keywords.cget("text"))
            self.transaction_full_details.append({"keyword_collection":self.keyword_collection,"source_reason":reason_source,"amount":amount})

    def delete_transation_box(self,frame):
        index = self.total_transation_frame.index(frame)
        self.total_transation_frame.pop(index)
        self.reason_added=len(self.total_transation_frame)
        self.transaction_full_details.pop(index)
        for i,each_frame in enumerate(self.total_transation_frame):
            label_1st= each_frame.winfo_children()[0]
            label_1st.configure(text=i+1)
        frame.destroy()

    def revert_all_transation_list(self):
        if self.total_transation_frame:
            for each_frame in self.total_transation_frame:
                each_frame.destroy()
            self.total_transation_frame=[]
            self.reason_added=0
            self.transaction_full_details=[]

    def Send_to_calculator(self):
        if not self.total_transation_frame:
            Notification(self.window,"Add Transation First plz")
        else:
            Transaction_book_calculator(self.transaction_full_details,self.transaction_name,self.add_or_sub,"trans",
                                        {"Func":self.Notifier_from_transation_calcu,"Revert":self.revert_all_transation_list})
            
    def Notifier_from_transation_calcu(self,text,color):
        fixed= (self.config.get("FILE_LOCATION","elements_for_money_track")).split(",")
        buttons_text=[]
        for each__ in self.keywordButtonlist:
                if each__.cget("text") in fixed:
                    pass
                else:
                    buttons_text.append(each__.cget("text"))
        newconfig=configparser.ConfigParser()
        newconfig.read(self.window.setting_ini_path)            
        newconfig.set("Keywords","Keywords",','.join(fixed + buttons_text))    
        with open(self.window.setting_ini_path,"w") as f:
                newconfig.write(f)
        Notification(self.window,text,color)
                
class Loan_BackEnd():
    def __init__(self,Window) :
        self.window= Window
        self.Main_frame=ctk.CTkFrame(self.window,fg_color="transparent")
        self.Main_frame.pack(expand=True,fill="both")

        ctk.CTkLabel(self.Main_frame,text="Money Manager",
                     fg_color="transparent",
                     font=ctk.CTkFont("Showcard Gothic Regular", 40, "bold")
                     ).place(relx=.15,rely=.0,relwidth=.7,relheight=.1)
        # back_btn
        ctk.CTkButton(self.Main_frame,text="↩",cursor="hand2",
                                font=ctk.CTkFont(family="Elephant",size=30),
                                fg_color="Blue",text_color="white",
                                command=self.GetBack,
                                ).place(relx=.04,rely=.02,relheight=.07,relwidth=.1)
        
        self.Menu_Frame=ctk.CTkTabview(self.Main_frame,
                                  segmented_button_selected_color="green",
                                  segmented_button_selected_hover_color="#86A789",
                                  segmented_button_unselected_color="#D2E3C8",
                                  text_color="black")
        self.Menu_Frame.place(relx=.05,relwidth=.9,rely=.09,relheight=.9)

        self.Menu_Frame.add("!!  Loan Given  !!")
        self.Menu_Frame.add("!!  Loan Taken  !!")
        self.degine_frame1=ctk.CTkFrame(self.Menu_Frame.tab("!!  Loan Given  !!"),fg_color="transparent")
        self.degine_frame1.pack(expand=True,fill="both")
        self.degine_frame2=ctk.CTkFrame(self.Menu_Frame.tab("!!  Loan Taken  !!"),fg_color="transparent")
        self.degine_frame2.pack(expand=True,fill="both")

        Loan_FrontEnd(self,self.degine_frame1,"Give")
        Loan_FrontEnd(self,self.degine_frame2,"Take")
    def GetBack(self):
        self.Main_frame.destroy()
        WinBackEnd(self.window)

class Loan_FrontEnd():
    def __init__(self,backend_cell,Design_frame,Give_Take):
        self.backend_cell=backend_cell
        self.window=backend_cell.window
        self.transaction_full_details=[]
        self.selected_Tangent_list=[]
        self.transaction_box_value=[]
        self.total_transation_frame=[]
        self.TangentButtonlist=[]
        self.reason_added=0
        self.Give_Take= Give_Take
        if self.Give_Take =="Take":
            self.conner_text="In Rs, [ Plus ➕ ]"
        else:
            self.conner_text="In Rs, [ Minus ➖ ]"

        ctk.CTkLabel(Design_frame,text="Loan's Transation",
                     font=ctk.CTkFont("Agency fb", 30, "bold"),
                     text_color="Blue",anchor="center", fg_color="transparent"
                     ).place(relx=0,rely=0,relheight=.1,relwidth=1)
        
        # Normal or Deposit Toggle Button
        self.Normal_deposit= ctk.CTkButton(Design_frame, text_color="Black",
                      text="Normal", cursor="hand2",
                      hover_color="green",
                      border_color="black",
                      fg_color="#C7C8CC",
                      border_width=2, 
                      command= self.Normal_or_Deposit
                      )
        self.Normal_deposit.place(relx=.02, rely=.012, relwidth=.2)
        self.N_D="N"
        self.due_date=datetime.now()+ timedelta(days=30)

        # due date 
        ctk.CTkLabel(backend_cell.Menu_Frame, text="Due Date:", font=("Helvetica", 14),fg_color="#C7C8CC", text_color="Black",corner_radius=2,
                     ).place(relx=.742,rely=.095,relwidth=.222, relheight= .033 )
        self.due_date_entry= DateEntry(Design_frame,
                year= int(self.due_date.strftime("%Y")),month=int(self.due_date.strftime("%m")), day=int(self.due_date.strftime("%d")),                        
                date_pattern='yyyy-mm-dd',background ="#1A5D1A",
                font=ctk.CTkFont("Agency fb", 20, "bold"),
                selectbackground='#00cc00', selectforeground='black',
                firstweekday="sunday",
                showothermonthdays =False,weekenddays=[7,7],
                normalbackground='white', normalforeground='black',
                weekendbackground='white', weekendforeground='black',
                headersbackground='black', headersforeground='white',
                )
        self.due_date_entry.bind("<<DateEntrySelected>>", self.reset_due_date)
        self.due_date_entry.place(relx=.75,rely=.04,relwidth=.23, relheight= .05 )

        self.config=configparser.ConfigParser()
        self.config.read(self.window.setting_ini_path)
        self.setting_Tangent_name=(self.config.get("Loan","tangents")).split(",")

        self.Upper_Half(Design_frame)
        self.Lower_Half(Design_frame)

    def reset_due_date(self,*args):
        self.due_date=self.due_date_entry.get_date()
    
    def Normal_or_Deposit(self):
        if self.Normal_deposit.cget("text") == "Normal":
            self.N_D="D"
            self.Normal_deposit.configure(text= "Deposit",border_color="White",fg_color="#7E2553",text_color="White")
        else:
            self.N_D="N"
            self.Normal_deposit.configure(text= "Normal",border_color="black",fg_color="#C7C8CC",text_color="Black") 

    def validate_amount(self, event):
        if event.char in ['', '\b','.']:return
        if not event.char.isdigit():
            self.amount_entry_box.bell()
            return 'break'

    def EnteR(self,*args):
        self.add_transation()
        pass

    def on_focus_in(self, *args):
        if self.amount_entry_box.get()==self.placeholder:
            self.amount_entry_box.delete(0, "end")      

    def handle(self, *args):
        pass        

    def Upper_Half(self,Design_frame):
        upper_half_frame=ctk.CTkFrame(Design_frame,corner_radius=0,)                             
        upper_half_frame.place(relx=0,rely=.1,relheight=.5,relwidth=1)

        # input main frame
        input_frame=ctk.CTkFrame(upper_half_frame,corner_radius=0)
        input_frame.place(relx=0,rely=.0,relwidth=1,relheight=.36)

        #Reason
        ctk.CTkLabel(input_frame,text="Reason").place(relx=.12,rely=.1)
        self.reason_entry_box=ctk.CTkEntry(input_frame,)
        self.reason_entry_box.place(relx=.25,rely=.1,relwidth=.6)

        #amount
        ctk.CTkLabel(input_frame,text="Amount :",).place(relx=.12,rely=.55)
        self.placeholder = self.conner_text
        self.amount_entry_box=ctk.CTkEntry(input_frame,)
        self.amount_entry_box.place(relx=.25,rely=.55,relwidth=.6)
        self.amount_entry_box.insert(0,self.placeholder)
        self.amount_entry_box.bind("<FocusIn>",self.on_focus_in)

        self.amount_entry_box.bind('<Key>', self.validate_amount)
        self.amount_entry_box.bind('<Tab>', self.handle)
        self.amount_entry_box.bind('<Control-v>', self.handle)
        self.amount_entry_box.bind('<Control-a>', self.handle)
        self.amount_entry_box.bind('<Control-c>', self.handle)
        self.amount_entry_box.bind('<Return>', self.EnteR)

        #Amount and reason adder button
        ctk.CTkButton(input_frame,text="\u2795 ",
                              fg_color="green",
                              hover_color="red",
                              cursor="hand2",
                              command=self.add_transation,
                              corner_radius=0
                              ).place(relx=.87,rely=.25,relheight=.5,relwidth=.1)

        #Tangent naming frame
        Tangent_frame=ctk.CTkFrame(upper_half_frame,corner_radius=0,)
        Tangent_frame.place(relx=0,rely=.36,relwidth=1,relheight=.64)

        #Tangent adder Frame
        Tangent_adder_frame=ctk.CTkFrame(Tangent_frame,corner_radius=0)
        Tangent_adder_frame.place(relx=.0,rely=.05,relwidth=1,relheight=.25)

        ctk.CTkLabel(Tangent_adder_frame,text="ADD Tangents:",
                     text_color="White",
                     fg_color="#2B2B2B",
                     corner_radius=0,
                     font=ctk.CTkFont("Algerian", 20, "bold")
                     ).place(relx=0,rely=.04,relwidth=.4)

        self.entrybox_Tangent_adder=ctk.CTkEntry(Tangent_adder_frame,)
        self.entrybox_Tangent_adder.place(relx=.39,rely=.04,relwidth=.5)

        self.entrybox_Tangent_adder.bind("<Return>",self.Tangent_adder)

        ctk.CTkButton(Tangent_adder_frame,text="+",fg_color="#864AF9",
                      command=self.Tangent_adder
                      ).place(relx=.9,rely=.04,relwidth=.07)

        #list of Tangent name frame
        self.list_of_Tangents_frame=ctk.CTkScrollableFrame(Tangent_frame, 
                                                      orientation="horizontal",
                                                      corner_radius=0,
                                                      )
        self.list_of_Tangents_frame.place(relx=0,rely=.25,relwidth=1,relheight=.75)

        self.list_1st_part=ctk.CTkFrame(self.list_of_Tangents_frame,)
        self.list_1st_part.pack(pady=8,padx=30)
        self.list_2nd_part=ctk.CTkFrame(self.list_of_Tangents_frame,)
        self.list_2nd_part.pack(pady=2,side="left",padx=30)

        #Tangent Name deleter/ Remover Box
        self.Tangent_Remover_btn=ctk.CTkButton(self.list_of_Tangents_frame,text="\U0001F5D1",
                      cursor="hand2",
                      hover_color="#FF004D",
                      command=self.remove_Tangent_btn,
                      width=20,
                      fg_color="red")
        
        if not self.setting_Tangent_name[0]:
            self.No_tangent_available=True
            self.setting_Tangent_name.pop(0)
            self.No_tangent_label= ctk.CTkLabel(Tangent_frame, 
                                                text="No Tangent Name Configured till Yet\nADD to configure",
                                                fg_color="#333333",corner_radius=2,text_color="White",
                                                font=ctk.CTkFont("Papyrus", 15, "bold"))
            self.No_tangent_label.place(relx=.1, rely=.34, relwidth=.8, relheight=.4)
        else:    
            self.No_tangent_available=False
            self.Tangent_Remover_btn.place(x=5,y=28)               
            self.Tangent_adder(Add_new=False)
        
    def Tangent_adder(self,Add_new=True): 
        if Add_new:
            new_Tangent=self.entrybox_Tangent_adder.get()
            if new_Tangent=='':
                Notification(self.window,"Write some text here First")
                return
            elif "," in new_Tangent :
                Notification(self.window,"PLz Dont keep \",\" in your Entry  field")
                return
            else:    
                for child in (self.list_1st_part.winfo_children()):
                    child.destroy()
                for child in (self.list_2nd_part.winfo_children()):
                    child.destroy() 
                self.TangentButtonlist=[]
                
                self.setting_Tangent_name.append(new_Tangent)
                self.entrybox_Tangent_adder.delete(0,'end')
                self.TangentsStructure_caller()    
                self.selected_Tangent_list=[]
        else:
            self.TangentsStructure_caller()      

    def TangentsStructure_caller(self):
        if self.No_tangent_available:
            self.No_tangent_label.place_forget()
            self.Tangent_Remover_btn.place(x=5,y=28)
            self.No_tangent_available=False
        for i in range(0,len(self.setting_Tangent_name),2):
            if i== len(self.setting_Tangent_name)-1:
                self.Tangents_structure(self.list_1st_part,self.setting_Tangent_name[i],i)
            else:    
                self.Tangents_structure(self.list_1st_part,self.setting_Tangent_name[i],i)
                self.Tangents_structure(self.list_2nd_part,self.setting_Tangent_name[i+1],i+1)
         
    def Tangents_structure(self,parent,text,index):
            Tangent_button=ctk.CTkButton(parent,
                          text=text,
                          fg_color="#4F6F52",
                          text_color="white",
                          hover_color="#739072",
                          cursor="hand2",
                          height=25,
                          command=lambda :self.select_Tangent(index)
                          )            
            Tangent_button.pack(padx=10,side="left")  
            self.TangentButtonlist.append(Tangent_button)

    def select_Tangent(self,index):
        key_word=self.TangentButtonlist[index]
        if key_word in self.selected_Tangent_list:
            key_word.configure(border_width=0,fg_color="#4F6F52")
            self.selected_Tangent_list.remove(key_word)
        else:
            self.selected_Tangent_list.append(key_word)
            key_word.configure(border_width=2,border_color="yellow",fg_color="GREEN")

    def remove_Tangent_btn(self,normal_call=True):
        if self.selected_Tangent_list:
            for each_btn in self.selected_Tangent_list:
                self.setting_Tangent_name.remove(each_btn.cget("text"))
                self.selected_Tangent_list.remove(each_btn)
                each_btn.destroy()
            self.remove_Tangent_btn(normal_call=False)
        else:
            if normal_call:
                Notification(self.window,'Plz select Tangent Name First and Hit Delete')     
   
    def Lower_Half(self,Design_frame):
        lower_half=ctk.CTkFrame(Design_frame,corner_radius=0,)                             
        lower_half.place(relx=0,rely=.6,relheight=.4,relwidth=1)

        self.list0fTransaction=ctk.CTkScrollableFrame(lower_half,corner_radius=0,)
        self.list0fTransaction.place(relx=0,rely=.05,relwidth=1,relheight=.82)

        title_frame=ctk.CTkFrame(self.list0fTransaction,height=20,corner_radius=0)
        title_frame.pack(side="top",fill="x")

        ctk.CTkLabel(title_frame,text="N0.").place(relx=0 ,rely=0,relwidth=.1,relheight=.6)
        ctk.CTkLabel(title_frame,text="Reason Here",).place(relx=.1,rely=0,relwidth=.75,relheight=.6)
        ctk.CTkLabel(title_frame,text="Amount").place(relx=.78,rely=0,relwidth=.12,relheight=.6)

        button_frame=ctk.CTkFrame(lower_half,corner_radius=0,)
        button_frame.place(relx=0,rely=.83,relwidth=1,relheight=.18)

        send_button  =ctk.CTkButton(button_frame,text="Send",fg_color="#4F6F52",
                                    command=self.Send_to_calculator,
                                    hover_color="#19A7CE",cursor="hand2")
        send_button  .place(relx=.25,rely=.0,relwidth=.2)
        Revert_button=ctk.CTkButton(button_frame,
                                    text="Revert",
                                    fg_color="#4F6F52",
                                    hover_color="#ED2B2A",
                                    command= self.revert_all_transation_list,
                                    cursor="hand2")
        Revert_button.place(relx=.55,rely=.0,relwidth=.2)       
        pass    

    def add_transation(self):
        valid_input = True
        
        if not self.reason_entry_box.get():
            Notification(self.window,f"Reason Needed to be Added !!!")
            valid_input = False
        elif not self.amount_entry_box.get() or self.amount_entry_box.get() == self.placeholder:
            Notification(self.window,"Amount Needed to be Added !!!")   
            valid_input = False
        elif self.No_tangent_available:
            Notification(self.window,"Plz First add your Tangent Name\nTo select just write and click Enter!!!")
            valid_input = False
        elif len(self.selected_Tangent_list) <1 or len(self.selected_Tangent_list)>1 :
            Notification(self.window,"You should Select only One Tangent Name !!!")
            valid_input = False 
        else:
            try:
                float(self.amount_entry_box.get())
            except:    
                Notification(self.window,f"In amount section String found, remove it !!!")
                valid_input = False

        if valid_input:
            self.reason_added+=1
            reason_source=self.reason_entry_box.get()
            amount=self.amount_entry_box.get()
            transaction_frame=ctk.CTkFrame(self.list0fTransaction,height=40)
            transaction_frame.pack(side="top",fill="x",pady=5,padx=5)

            ctk.CTkLabel(transaction_frame,text=self.reason_added).place(relx=0 ,rely=0.2,relwidth=.05,relheight=.6)
            ctk.CTkLabel(transaction_frame,text=reason_source).place(relx=.05,rely=0.2,relwidth=.72,relheight=.6)
            ctk.CTkLabel(transaction_frame,text=amount,fg_color="black").place(relx=.78,rely=0.2,relwidth=.14,relheight=.6)
            ctk.CTkButton(transaction_frame,
                          text="\u2796 ",
                          fg_color="red",
                          cursor="hand2",
                          hover_color="#EC255A",
                          corner_radius=0,
                          font=("Arial", 18,"bold"),
                          command=lambda index=self.reason_added,f=transaction_frame:self.delete_transation_box(f)
                          ).place(relx=.93,rely=0.2,relwidth=.06,relheight=.6)  
            ctk.CTkLabel(transaction_frame,text=self.N_D,fg_color="blue").place(relx=.05 ,rely=0.2,relwidth=.04,relheight=.6)
            
            self.amount_entry_box.delete(0,"end")
            self.total_transation_frame.append(transaction_frame)
            self.transaction_full_details.append({"Due_date": self.due_date.strftime("%Y-%m-%d"),
                                                  "Tangent_Name":[self.selected_Tangent_list[0].cget("text")],
                                                  "source_reason":reason_source,
                                                  "amount":amount, 
                                                  "deposit_normal":self.N_D})

    def delete_transation_box(self,frame):
        index = self.total_transation_frame.index(frame)
        self.total_transation_frame.pop(index)
        self.reason_added=len(self.total_transation_frame)
        self.transaction_full_details.pop(index)
        for i,each_frame in enumerate(self.total_transation_frame):
            label_1st= each_frame.winfo_children()[0]
            label_1st.configure(text=i+1)
        frame.destroy()

    def revert_all_transation_list(self):
        if self.total_transation_frame:
            for each_frame in self.total_transation_frame:
                each_frame.destroy()
            self.total_transation_frame=[]
            self.reason_added=0
            self.transaction_full_details=[]

    def Send_to_calculator(self):
        if not self.total_transation_frame:
            Notification(self.window,"Add Transation First plz")
        else:
            Transaction_book_calculator(self.transaction_full_details,"Loan",self.Give_Take,"Loan",
                                        {"Func":self.Notifier_from_transation_calcu,"Revert":self.revert_all_transation_list})
            pass
            
    def Notifier_from_transation_calcu(self,text,color):
        fixed= (self.config.get("Loan","tangents")).split(",")
        if fixed[0] =="":
            fixed.pop(0)
        buttons_text=[]
        for each_tangent in self.setting_Tangent_name:
                if each_tangent in fixed:
                    pass
                else:
                    buttons_text.append(each_tangent)
        newconfig=configparser.ConfigParser()
        newconfig.read(self.window.setting_ini_path)            
        newconfig.set("Loan","tangents",','.join(fixed + buttons_text))    
        with open(self.window.setting_ini_path,"w") as f:
                newconfig.write(f)
        Notification(self.window,text,color)

class Notification(ctk.CTkFrame):
    def __init__(self,window, text, color="red", delay_time=5):
        super().__init__(master=window,fg_color=color,corner_radius=0)
        ctk.CTkLabel(self,text=text,).pack(anchor="center",pady=8)
        ctk.CTkButton(self,text="✘",
                      text_color="white",
                      fg_color="black",
                      corner_radius=0,
                      hover_color="red",
                      cursor="hand2",
                      command=self.destroy_bar
                      ).place(relx=.94,rely=0,relwidth= .06,relheight=.4)
        delay_time=delay_time*1000
        self.place(relx=.1,rely=.6,relwidth=.8)
        window.bell()

        window.after(delay_time,self.destroy_bar)

    def destroy_bar(self):self.destroy()



class CommandLIne():
    def __init__(self, path):
        self.settingpath=path
        self.config=configparser.ConfigParser()
        self.config.read(self.settingpath)
        self.elements=(self.config.get("FILE_LOCATION","elements_for_money_track")).split(",")
        self.keywords=(self.config.get("Keywords","keywords")).split(",")
        self.tangentsL=(self.config.get("Loan","tangents")).split(",")
        self.added_elements= self.elements + ['Loan','Check Money']
        Question_to_ask='\n\n'
        for index,each_elem in enumerate(self.added_elements):
            Question_to_ask= Question_to_ask+f"{index+1}) {each_elem}\t\t"

        Question_to_ask= Question_to_ask+ f"\n\n Where You Want to Go, hint: (choose 1 to {len(self.added_elements)}):\t"

        while True:
            try:
                Answer_to_go= int(input(Question_to_ask))
                if (Answer_to_go) >0 and (Answer_to_go)<=len(self.added_elements):break         
                else:raise ValueError("error")            
            except ValueError:
                print(f"\nsorry wrong input, Hint: (Give number between 1 to {len(self.added_elements)})")
                print("\t\tPlz Try again")  

        if Answer_to_go <= len(self.elements):
            self.Do_trans_on(Answer_to_go-1)  
        elif  self.added_elements[Answer_to_go-1]== "Loan":
            self.Loan_section()
        else:
            self.Check_see()  

        change_ui=input("\n(you change Ui mode by tryping \"change\")\nPress Enter key to End:")
        if change_ui.lower() == "change":
            config=configparser.ConfigParser()
            config.read(path)
            config.set("Ui_mode","ui_mode","True")
            with open(path,'w') as file:
                config.write(file)
            print("Truned Ui Mode Onn")

    def Check_see(self):
        print("Sir your money are:\n\n")
        print("_____________________________________________")
        print("sn.|  Inside      |        Money")
        print("---------------------------------------------")
        for index,each_elem in enumerate(self.elements):
            solid_name= "_".join(each_elem.split(" "))
            Amt_s=(self.config.get(solid_name,f"{solid_name}_amounts"))
            print(f"{index+1})| {each_elem}  |        {Amt_s}")
            print("_____________________-_____________________")

        print("\n\n\n") 
        print("Loan sections ") 
        print("__________________________________________________")
        print("sn. | Tangent Name | Amt To Take | Amt to Give ")  
        print("------------------------------------------------")
        if self.tangentsL[0] !="":
            for index, eachTangent in enumerate(self.tangentsL):
                try:
                    solid_name="_".join(eachTangent.split(" "))+"_amt"
                    amots=float((self.config.get("Loan",solid_name)))
                    if amots >=0:
                        take_a= amots
                        give_a=0 
                    else:
                        take_a= 0
                        give_a=amots 

                    print(f"{index+1})| {eachTangent}  | {take_a} |{give_a}")
                    print("_________________-_____________________________________")
                except:pass    
        else:
            print(" Sorry No data Feeded yet")
            print("_________________-_____________________________________")

    def Loan_section(self):
        print(f"Your Transaction on Loan")
        L_t_deatils=[]
        num_of_trans=0
        while True:
            reason = (input("\nEnter your reason:\t"))
            if reason.lower() == "end":
                break
            elif reason.lower()=="ed" and num_of_trans>=1:
                num_of_trans = num_of_trans-1
                L_t_deatils.pop(num_of_trans)
            else:
                #amt
                while True:    
                    try:
                        money = float(input("\nEnter the money [to give -, to take +]:\tRs. "))
                        break
                    except:    
                        print("Sorry you have to enter money in number, Try again")
                
                # tangent
                while True:
                    if self.tangentsL[0]!='':
                        full_text = ''.join([f"{index+1}) {each_tangent}\t\t" for index, each_tangent in enumerate(self.tangentsL)])
                        tangents_select=(input(f"\n {full_text}\nEnter number,(select only one Tangent), to ADD press add: \t"))
                        if tangents_select.lower() == "add":
                            new_tangent= input("\n Enter New Keyword:\t")
                            if "," in new_tangent:
                                print("plz dont put comma here, Reverted")
                            else:    
                                self.tangentsL.append(new_tangent)
                        else:
                            try:
                                if int(tangents_select) < len(self.tangentsL)+1 and int(tangents_select) !=0:
                                    seleced_tangent = self.tangentsL[int(tangents_select)-1]
                                    break
                                else:
                                    raise ValueError("error")
                            except :
                                print("Sorry You have not entered the number as requested")  
                    else:
                        print("\n\nNo TAagent Added PLz Add new one")
                        new_tangent= input("Enter New Keyword:")
                        if "," in new_tangent:
                            print("plz dont put comma here, Reverted")
                        else:    
                            self.tangentsL[0]=new_tangent
            
                # due date and normal 0r deposite
                while True:
                    d_d=(datetime.now()+ timedelta(days=30)).strftime("%Y-%m-%d")
                    due_date = input(f"Change Due date for this \'{d_d}\' or just press enter:\t")
                    if due_date =="":break
                    else:
                        try:
                            splited= due_date.split("-")
                            d_d=datetime(int(splited[0]),int(splited[1]),int(splited[2]))
                            break
                        except:pass    
                while True:
                    n_d="N"
                    normal_deposit= input("Is your Trans Normal then press enter or else press D:\t")
                    if normal_deposit == "":break    
                    elif normal_deposit =="D":n_d="D"

                L_t_deatils.append([seleced_tangent,reason,money, d_d,n_d])
                num_of_trans+=1
        Give_trans=[]
        Take_trans=[]
                
        for each_trans in L_t_deatils:
            if float(each_trans[2]) >=0:
                Take_trans.append({"Due_date":each_trans[3],"Tangent_Name":[each_trans[0]],"source_reason":each_trans[1],"amount":float(each_trans[2]),"deposit_normal":each_trans[4]})
            else:
                Give_trans.append({"Due_date":each_trans[3],"Tangent_Name":[each_trans[0]],"source_reason":each_trans[1],"amount":float(each_trans[2]),"deposit_normal":each_trans[4]})
        if Take_trans:        
            Transaction_book_calculator(Take_trans,"Loan","Take","Loan",)
        if Give_trans:
            Transaction_book_calculator(Give_trans,"Loan","Give","Loan",)

        newconfig=configparser.ConfigParser()
        newconfig.read(self.settingpath)            
        key_txt= ','.join(self.tangentsL)    
        newconfig.set("Loan","tangents",key_txt)
        with open(self.settingpath,'w') as f:
            newconfig.write(f)

    def Do_trans_on(self,index_num):
        print(f"Your Transaction on {self.elements[index_num]}")
        trans_details=[]
        num_of_trans=0
        while True:
            reason = (input("\nEnter your reason:\t"))
            if reason.lower() == "end":
                break
            elif reason.lower()=="ed" and num_of_trans>=1:
                num_of_trans = num_of_trans-1
                trans_details.pop(num_of_trans)
            else:
                while True:    
                    try:
                        money = float(input("\nEnter the money:\tRs. "))
                        break
                    except:    
                        print("Sorry you have to enter money in number, Try again")
                while True:
                    full_text = ''.join([f"{index+1}) {each_keywords}\t\t" for index, each_keywords in enumerate(self.keywords)])
                    keywords_select=input(f"\n {full_text}\nEnter number using comma,(select atleast 2 keywords like (2,3)), to ADD press add: \t")
                    if keywords_select.lower() == "add":
                        new_keyword= input("\n Enter New Keyword:")
                        if "," in new_keyword:
                            print("plz dont put comma here, Reverted")
                        else:    
                            self.keywords.append(new_keyword)
                    else:
                        try:
                            seleced_keywords =[self.keywords[int(index)-1] for index in keywords_select.split(',')]
                            break
                        except :
                            print("Sorry You have not entered the number as demanded")
                trans_details.append([seleced_keywords,reason,money])
                num_of_trans+=1
        ADD_trans=[]
        SUB_trans=[]
        
        for each_trans in trans_details:
            if float(each_trans[2]) >=0:
                ADD_trans.append({"keyword_collection":each_trans[0],"source_reason":each_trans[1],"amount":float(each_trans[2])})
            else:
                SUB_trans.append({"keyword_collection":each_trans[0],"source_reason":each_trans[1],"amount":abs(each_trans[2])})
        if SUB_trans:        
            Transaction_book_calculator(SUB_trans,self.elements[index_num],"SUB","trans",)
        if ADD_trans:
            Transaction_book_calculator(ADD_trans,self.elements[index_num],"ADD","trans",)

        key_txt= ','.join(self.keywords)    
        newconfig=configparser.ConfigParser()
        newconfig.read(self.settingpath)            
        newconfig.set("Keywords","keywords",key_txt)
        with open(self.settingpath,'w') as f:
            newconfig.write(f)    
    

if __name__=="__main__":    
    path=f"{(os.path.dirname(os.path.abspath(__file__)))}\\setting.ini"
    if not os.path.exists(path):
        with open(path,'w') as file:
            config=configparser.ConfigParser()
            config.read(path)
            config.add_section('Ui_mode')
            config.set("Ui_mode","ui_mode","True")
            config.write(file)
            ui_mode="True"
    else:
        config=configparser.ConfigParser()
        config.read(path)
        ui_mode= config.get("Ui_mode","ui_mode")
    if ui_mode == 'True':
        Window()
    else:
        CommandLIne(path)
        pass    











